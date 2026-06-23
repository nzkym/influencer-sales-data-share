"""
파마브로스 판매현황 파일공유 모듈
─────────────────────────────────────────
- 네이버 API로 개별 주문 데이터 수집
- 엑셀(.xlsx) 파일 생성
- 구글 드라이브 공유 폴더에 업로드

실행 스케줄 (한국시간 기준):
  시작일        : 14시, 16시  → 당일 데이터
  D+1 ~ 종료일  : 매일 10시   → 시작일~당일 누적 데이터
  종료일 다음날  : 10시        → 시작일~종료일 최종 전체 데이터
"""

import os
import io
import re
from datetime import datetime, timezone, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from google.oauth2.credentials import Credentials as OAuthCredentials
from google.auth.transport.requests import Request as AuthRequest
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload

# ── 한국시간 ──────────────────────────────────────────────
KST = timezone(timedelta(hours=9))


# ── 영업일 계산 (한국 공휴일 + 주말 제외) ─────────────────
def _business_days_after(start, n: int):
    """start 날짜로부터 한국 공휴일·주말 제외 n 영업일 후 날짜 반환."""
    from datetime import date as _date
    try:
        import holidays as _hols
        kr = _hols.KR(years=range(start.year, start.year + 2))
    except ImportError:
        kr = {}
    current = start
    count = 0
    while count < n:
        current += timedelta(days=1)
        if current.weekday() < 5 and current not in kr:   # 평일 + 비공휴일
            count += 1
    return current


def get_delete_date(end_date_str: str):
    """최종 파일 자동삭제 날짜 반환 (최종 업로드일로부터 3 영업일 후).
    최종 업로드일 = end_date + 1일.
    """
    end        = _parse_date(end_date_str)
    final_day  = end + timedelta(days=1)   # 종료일 다음날 = 최종 업로드일
    return _business_days_after(final_day, 3)


def now_kst() -> datetime:
    return datetime.now(KST)


# ── 실행 여부 판단 ────────────────────────────────────────
def should_run(start_date_str: str, end_date_str: str) -> tuple[bool, bool]:
    """
    현재 KST 시각 기준으로 파마브로스 업로드를 실행해야 하는지 판단.

    반환: (should_run: bool, is_final: bool)
      is_final=True  → 종료일 다음날 최종 업로드
      is_final=False → 중간 업로드
    """
    from datetime import date as _date
    now   = now_kst()
    today = now.date()
    hour  = now.hour

    start = _parse_date(start_date_str)
    end   = _parse_date(end_date_str)

    minute = now.minute

    # 종료일 다음날 10시: 최종 업로드 (10:00~10:14만 실행, 3회 시도)
    if today == end + timedelta(days=1) and hour == 10 and minute < 15:
        return True, True

    # 시작일 14시 또는 16시 (첫 15분만 실행, 3회 시도)
    if today == start and hour in (14, 16) and minute < 15:
        return True, False

    # D+1 이후 종료일까지 매일 10시 (10:00~10:14만 실행, 3회 시도)
    if start < today <= end and hour == 10 and minute < 15:
        return True, False

    return False, False


def _parse_date(s: str):
    from datetime import date
    parts = re.split(r"[.\-/]", s.strip())
    return date(int(parts[0]), int(parts[1]), int(parts[2]))


# ── 엑셀 파일 생성 ────────────────────────────────────────
def _col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


def create_excel(
    orders: list,
    title: str,
    date_from: str,
    date_to: str,
    is_final: bool,
    product_url: str = "",
) -> bytes:
    """
    주문 데이터를 받아 엑셀 바이트를 반환합니다.

    orders: [{"주문번호", "주문일시", "주문상태", "옵션", "주문수량"}, ...]
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "판매현황"

    # ── 색상 정의 ─────────────────────────────────────────
    HEADER_FILL  = PatternFill("solid", fgColor="1A2744")   # 진남색
    TITLE_FILL   = PatternFill("solid", fgColor="2D3F6B")   # 중간남색
    TOTAL_FILL   = PatternFill("solid", fgColor="EEF1FA")   # 연보라
    WHITE_FONT   = Font(name="맑은 고딕", color="FFFFFF", bold=True, size=11)
    BODY_FONT    = Font(name="맑은 고딕", size=10)
    BOLD_FONT    = Font(name="맑은 고딕", bold=True, size=10)
    TOTAL_FONT   = Font(name="맑은 고딕", bold=True, size=10, color="1A2744")
    CENTER       = Alignment(horizontal="center", vertical="center")
    LEFT         = Alignment(horizontal="left",   vertical="center")
    thin         = Side(style="thin", color="CCCCCC")
    BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── 1행: 제목 ─────────────────────────────────────────
    label = f"{'[최종]' if is_final else '[중간]'} {title} 판매현황   ({date_from} ~ {date_to})"
    ws.merge_cells("A1:E1")
    cell = ws["A1"]
    cell.value = label
    cell.font  = Font(name="맑은 고딕", color="FFFFFF", bold=True, size=13)
    cell.fill  = TITLE_FILL
    cell.alignment = CENTER
    ws.row_dimensions[1].height = 32

    # ── 2행: 발행 시각 ────────────────────────────────────
    ws.merge_cells("A2:E2")
    ts_cell = ws["A2"]
    ts_cell.value     = f"업로드: {now_kst().strftime('%Y-%m-%d %H:%M')} KST    총 {len(orders)}건"
    ts_cell.font      = Font(name="맑은 고딕", size=9, color="888888")
    ts_cell.alignment = RIGHT = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── 3행: 상품링크 ────────────────────────────────────
    LINK_FILL = PatternFill("solid", fgColor="F0F4FF")
    ws.merge_cells("A3:E3")
    lk = ws["A3"]
    lk.value     = f"상품링크: {product_url}" if product_url else "상품링크: -"
    lk.font      = Font(name="맑은 고딕", size=9, color="1A2744")
    lk.fill      = LINK_FILL
    lk.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 16

    # ── 4행: 헤더 ─────────────────────────────────────────
    headers = ["주문번호", "주문일시", "주문상태", "옵션", "주문수량", "단가", "결제금액"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font      = WHITE_FONT
        c.fill      = HEADER_FILL
        c.alignment = CENTER
        c.border    = BORDER
    ws.row_dimensions[4].height = 22

    # ── 데이터 행 ─────────────────────────────────────────
    EVEN_FILL = PatternFill("solid", fgColor="F8F9FB")
    for idx, order in enumerate(orders, 1):
        row = idx + 4
        row_fill = EVEN_FILL if idx % 2 == 0 else None
        unit_price = int(order.get("단가", 0))
        qty = int(order.get("주문수량", 0))
        values = [
            order.get("주문번호", ""),
            order.get("주문일시", ""),
            order.get("주문상태", ""),
            order.get("옵션", ""),
            qty,
            unit_price,
            unit_price * qty,
        ]
        aligns = [CENTER, CENTER, CENTER, LEFT, CENTER, CENTER, CENTER]
        for col, (val, aln) in enumerate(zip(values, aligns), 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = BODY_FONT
            c.alignment = aln
            c.border    = BORDER
            if row_fill:
                c.fill = row_fill
        ws.row_dimensions[row].height = 18

    # ── 합계 행 ───────────────────────────────────────────
    total_row = len(orders) + 5
    total_qty = sum(int(o.get("주문수량", 0)) for o in orders)
    total_amt = sum(int(o.get("단가", 0)) * int(o.get("주문수량", 0)) for o in orders)
    ws.merge_cells(f"A{total_row}:D{total_row}")
    tc1 = ws[f"A{total_row}"]
    tc1.value     = "총 주문수량"
    tc1.font      = TOTAL_FONT
    tc1.fill      = TOTAL_FILL
    tc1.alignment = CENTER
    tc1.border    = BORDER
    tc2 = ws.cell(row=total_row, column=5, value=total_qty)
    tc2.font      = TOTAL_FONT
    tc2.fill      = TOTAL_FILL
    tc2.alignment = CENTER
    tc2.border    = BORDER
    ws.cell(row=total_row, column=6).fill = TOTAL_FILL
    ws.cell(row=total_row, column=6).border = BORDER
    tc3 = ws.cell(row=total_row, column=7, value=total_amt)
    tc3.font      = TOTAL_FONT
    tc3.fill      = TOTAL_FILL
    tc3.alignment = CENTER
    tc3.border    = BORDER
    ws.row_dimensions[total_row].height = 22

    # ── 열 너비 ───────────────────────────────────────────
    _col_width(ws, 1, 22)   # 주문번호
    _col_width(ws, 2, 20)   # 주문일시
    _col_width(ws, 3, 14)   # 주문상태
    _col_width(ws, 4, 36)   # 옵션
    _col_width(ws, 5, 10)   # 주문수량
    _col_width(ws, 6, 12)   # 단가
    _col_width(ws, 7, 12)   # 결제금액

    # ── 틀 고정 (4행 헤더 이후) ──────────────────────────
    ws.freeze_panes = "A5"

    # 바이트로 반환
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Google Drive 업로드 (OAuth2 — 사장님 계정 권한) ────────
def _drive_service_oauth(client_id: str, client_secret: str, refresh_token: str):
    """사장님 구글 계정 OAuth2 토큰으로 Drive 서비스 생성."""
    creds = OAuthCredentials(
        token=None,
        refresh_token=refresh_token,
        client_id=client_id,
        client_secret=client_secret,
        token_uri="https://oauth2.googleapis.com/token",
    )
    creds.refresh(AuthRequest())
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def upload_to_drive(
    client_id: str,
    client_secret: str,
    refresh_token: str,
    folder_id: str,
    file_bytes: bytes,
    file_name: str,
    title: str = "",
) -> str:
    """
    사장님 구글 드라이브 폴더에 엑셀 파일 업로드.
    같은 캠페인(title_ 로 시작하는 파일)은 업로드 전에 모두 삭제.
    """
    service = _drive_service_oauth(client_id, client_secret, refresh_token)
    media   = MediaIoBaseUpload(
        io.BytesIO(file_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # 같은 캠페인 기존 파일 전부 삭제 (title_ 로 시작하는 파일 모두)
    if title:
        safe_prefix = re.sub(r'[\\/:*?"<>|]', "_", title).strip() + "_"
        all_in_folder = service.files().list(
            q=f"'{folder_id}' in parents and trashed=false",
            fields="files(id,name)",
        ).execute().get("files", [])
        for f in all_in_folder:
            if f["name"].startswith(safe_prefix):
                service.files().delete(fileId=f["id"]).execute()
                print(f"  [Drive] 기존 파일 삭제: {f['name']}")

    # 새 파일 업로드
    uploaded = service.files().create(
        body={"name": file_name, "parents": [folder_id]},
        media_body=media,
        fields="id",
    ).execute()
    file_id = uploaded["id"]
    print(f"  [Drive] 업로드 완료: {file_name}")

    return f"https://drive.google.com/file/d/{file_id}/view"


# ── 파일명 생성 ───────────────────────────────────────────
def make_filenames(title: str, is_final: bool, date_from: str = "", date_to: str = "") -> list[str]:
    """
    업로드할 파일명 반환 (항상 1개).

    정기 업로드: 힐링희_5월20일_14시_업로드건.xlsx
    최종 업로드: 힐링희_5월25일_10시_최종_업로드건.xlsx
    """
    now  = now_kst()
    safe = re.sub(r'[\\/:*?"<>|]', "_", title).strip()
    ts   = f"{now.month}월{now.day}일_{now.hour}시"

    if is_final:
        return [f"{safe}_{ts}_최종_업로드건.xlsx"]
    else:
        return [f"{safe}_{ts}_업로드건.xlsx"]


# ── 메인 실행 함수 ────────────────────────────────────────
def run_pharmabros(
    campaign: dict,
    credentials_path: str,
    oauth_client_id: str,
    oauth_client_secret: str,
    oauth_refresh_token: str,
    drive_folder_id: str,
):
    """
    파마브로스 판매현황 구글 드라이브 업로드 실행 (사장님 계정 OAuth2).
    campaign 키: title, product_no, date_from, date_to, api_id, api_secret
    """
    import naver_api

    title     = campaign["title"]
    date_from = campaign["date_from"]
    date_to   = campaign["date_to"]
    is_final  = campaign.get("is_final", False)

    # 실제 조회 기간
    if is_final:
        query_to = date_to
    else:
        today_str = now_kst().strftime("%Y-%m-%d")
        query_to  = min(date_to, today_str)

    print(f"\n  [파마브로스] '{title}' 데이터 수집 시작")
    print(f"  {'[최종]' if is_final else '[중간]'} 기간: {date_from} ~ {query_to}")

    orders = naver_api.get_pharmabros_orders(
        client_id=campaign["api_id"],
        client_secret=campaign["api_secret"],
        product_no=campaign["product_no"],
        date_from=date_from,
        date_to=query_to,
    )

    # 엑셀 생성
    product_url = campaign.get("url", "").split("?")[0]  # 쿼리스트링 제거
    excel_bytes = create_excel(
        orders=orders,
        title=title,
        date_from=date_from,
        date_to=query_to,
        is_final=is_final,
        product_url=product_url,
    )

    # 파일명: 제목_시작일~종료일_시각.xlsx
    file_name = make_filenames(title, is_final, date_from, query_to)[0]

    # 구글 드라이브 업로드 (사장님 계정) — 기존 파일 삭제 후 신규 업로드
    file_url = upload_to_drive(
        client_id=oauth_client_id,
        client_secret=oauth_client_secret,
        refresh_token=oauth_refresh_token,
        folder_id=drive_folder_id,
        file_bytes=excel_bytes,
        file_name=file_name,
        title=title,
    )

    folder_url = f"https://drive.google.com/drive/folders/{drive_folder_id}"
    print(f"  [파마브로스] ✅ 업로드 완료: {file_name}")
    print(f"  [파마브로스] 폴더: {folder_url}")
    return folder_url, [(file_name, file_url)]


def read_xlsx_rows_from_drive(
    client_id: str,
    client_secret: str,
    refresh_token: str,
    folder_id: str,
    title: str,
) -> list:
    """드라이브 폴더(+완료 하위폴더)에서 해당 캠페인 xlsx를 읽어 전체 행 반환.
    반환: [{"주문번호":..., "주문일시":..., "주문상태":..., "옵션":..., "주문수량":..., "단가":..., "결제금액":...}, ...]
    """
    import io
    from openpyxl import load_workbook

    service     = _drive_service_oauth(client_id, client_secret, refresh_token)
    safe_prefix = re.sub(r'[\\/:*?"<>|]', "_", title).strip() + "_"

    # 메인 폴더 + 완료 폴더 모두 검색
    search_folders = [folder_id]
    try:
        done_files = service.files().list(
            q=f"'{folder_id}' in parents and name='완료' and mimeType='application/vnd.google-apps.folder' and trashed=false",
            fields="files(id)",
        ).execute().get("files", [])
        if done_files:
            search_folders.append(done_files[0]["id"])
    except Exception:
        pass

    target = None
    for fid in search_folders:
        files = service.files().list(
            q=f"'{fid}' in parents and trashed=false",
            fields="files(id,name,createdTime)",
            orderBy="createdTime desc",
        ).execute().get("files", [])
        for f in files:
            if f["name"].startswith(safe_prefix) and f["name"].endswith(".xlsx"):
                target = f
                break
        if target:
            break

    if not target:
        print(f"  [Drive] '{title}' 파일 없음")
        return []

    print(f"  [Drive] 읽기: {target['name']}")
    content = service.files().get_media(fileId=target["id"]).execute()
    wb = load_workbook(filename=io.BytesIO(content), read_only=True)
    ws = wb.active

    # 헤더 찾기 (4행이 헤더)
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        cells = [str(c or "").strip() for c in row]
        if "주문번호" in cells:
            header_row = cells
            break
    if not header_row:
        wb.close()
        return []

    header_start = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        cells = [str(c or "").strip() for c in row]
        if "주문번호" in cells:
            header_start = i
            break

    result = []
    for row in ws.iter_rows(min_row=header_start + 1, values_only=True):
        if not row or not row[0]:
            continue
        entry = {}
        for i, h in enumerate(header_row):
            if i < len(row) and h:
                entry[h] = row[i] if row[i] is not None else ""
        if entry.get("주문번호"):
            result.append(entry)

    wb.close()
    print(f"  [Drive] {len(result)}건 읽기 완료")
    return result


def _get_or_create_done_folder(service, parent_folder_id: str) -> str:
    """'완료' 하위 폴더 ID 반환. 없으면 생성."""
    results = service.files().list(
        q=f"'{parent_folder_id}' in parents and name='완료' and mimeType='application/vnd.google-apps.folder' and trashed=false",
        fields="files(id)",
    ).execute().get("files", [])
    if results:
        return results[0]["id"]
    folder = service.files().create(body={
        "name": "완료",
        "mimeType": "application/vnd.google-apps.folder",
        "parents": [parent_folder_id],
    }, fields="id").execute()
    print(f"  [Drive] '완료' 폴더 생성")
    return folder["id"]


def delete_campaign_files(
    client_id: str,
    client_secret: str,
    refresh_token: str,
    folder_id: str,
    title: str,
) -> list:
    """드라이브 폴더에서 해당 캠페인 파일을 '완료' 폴더로 이동. 이동된 파일명 목록 반환."""
    service     = _drive_service_oauth(client_id, client_secret, refresh_token)
    safe_prefix = re.sub(r'[\\/:*?"<>|]', "_", title).strip() + "_"

    all_files = service.files().list(
        q=f"'{folder_id}' in parents and trashed=false",
        fields="files(id,name)",
    ).execute().get("files", [])

    moved = []
    done_folder_id = None
    for f in all_files:
        if f["name"].startswith(safe_prefix):
            if not done_folder_id:
                done_folder_id = _get_or_create_done_folder(service, folder_id)
            service.files().update(
                fileId=f["id"],
                addParents=done_folder_id,
                removeParents=folder_id,
            ).execute()
            moved.append(f["name"])
            print(f"  [Drive] 완료 폴더로 이동: {f['name']}")

    return moved
