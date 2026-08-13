"""
인플루언서 판매 데이터 공유 프로그램
- 캠페인 목록: 구글 시트(MASTER_SHEET_URL)에서 읽기
- 네이버 스마트스토어 API로 판매 데이터 조회
- 각 인플루언서 구글 시트에 결과 기록

실행 방법:
  python main.py          → 로컬 PC: 시작 즉시 1회 + 3시간마다 반복
  python main.py --once   → GitHub Actions: 1회 실행 후 종료
"""

import re
import os
import sys
import time
import schedule
import requests as _requests
from datetime import datetime, date, timezone, timedelta
from dotenv import load_dotenv
from pathlib import Path

import json
import gspread
from google.oauth2.service_account import Credentials

import naver_api
import sheets
import pharmabros

# .env 로딩
BASE_DIR = Path(__file__).parent
load_dotenv(BASE_DIR / ".env")

MASTER_SHEET_URL = os.getenv("MASTER_SHEET_URL")

# 스토어명 → API 키 매핑
STORE_CREDENTIALS = {
    "nutone":   (os.getenv("NUTONE_CLIENT_ID"),   os.getenv("NUTONE_CLIENT_SECRET")),
    "jdhealth": (os.getenv("JDHEALTH_CLIENT_ID"), os.getenv("JDHEALTH_CLIENT_SECRET")),
    "nutpet":   (os.getenv("NUTPET_CLIENT_ID"),   os.getenv("NUTPET_CLIENT_SECRET")),
    "nutmedi":  (os.getenv("NUTMEDI_CLIENT_ID"),  os.getenv("NUTMEDI_CLIENT_SECRET")),
}

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
TELEGRAM_CHAT_ID   = os.getenv("TELEGRAM_CHAT_ID")

# 공구수수료 별도 시트 (fill_lmno.py와 동일)
COMM_SHEET_URL = os.getenv(
    "COMM_SHEET_URL",
    "https://docs.google.com/spreadsheets/d/1tNYLgKB-rXcwF5ZdjxT3xti4ygvNl6SEJ-x0w8Gmjt0/edit",
)

CREDENTIALS_PATH = str(BASE_DIR / "credentials" / "google-credentials.json")
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

SOLDOUT_NOTIFIED_FILE = BASE_DIR / "soldout_notified.json"
PHARMABROS_PRICES_DIR = BASE_DIR / "pharmabros_prices"

# 파마브로스 파일공유 — 사장님 구글 계정 OAuth2 + Drive 폴더
PHARMABROS_OAUTH_CLIENT_ID         = os.getenv("OAUTH_CLIENT_ID", "")
PHARMABROS_OAUTH_CLIENT_SECRET     = os.getenv("OAUTH_CLIENT_SECRET", "")
PHARMABROS_OAUTH_REFRESH_TOKEN     = os.getenv("OAUTH_REFRESH_TOKEN", "")
PHARMABROS_DRIVE_FOLDER_ID         = os.getenv("PHARMABROS_DRIVE_FOLDER_ID", "")
PHARMABROS_SETTLEMENT_FOLDER_ID    = os.getenv("PHARMABROS_SETTLEMENT_FOLDER_ID", "")

# 파마브로스파일공유 판별 함수 (띄어쓰기 무관)
def _is_pharmabros(value: str) -> bool:
    """K열 값이 '파마브로스파일공유'이면 True (공백 무시)."""
    return re.sub(r"\s+", "", str(value)) == "파마브로스파일공유"


def _load_soldout_notified() -> set:
    """품절 알림을 이미 보낸 상품번호 목록 로드."""
    try:
        if SOLDOUT_NOTIFIED_FILE.exists():
            return set(json.loads(SOLDOUT_NOTIFIED_FILE.read_text(encoding="utf-8")))
    except Exception:
        pass
    return set()


def _save_soldout_notified(notified: set):
    SOLDOUT_NOTIFIED_FILE.write_text(json.dumps(list(notified)), encoding="utf-8")




# ── 텔레그램 알림 ────────────────────────────────────────
def send_telegram(message: str):
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        return
    try:
        _requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage",
            data={"chat_id": TELEGRAM_CHAT_ID, "text": message},
            timeout=10,
        )
    except Exception:
        pass


# ── 날짜 파싱 ────────────────────────────────────────────
def parse_date(date_str: str) -> date:
    date_str = date_str.strip()
    parts = re.split(r"[.\-/]", date_str)
    if len(parts) == 3:
        return date(int(parts[0]), int(parts[1]), int(parts[2]))
    raise ValueError(f"날짜 형식을 인식할 수 없습니다: {date_str}")


def extract_product_no(url: str) -> str:
    match = re.search(r"/products/(\d+)", url)
    if not match:
        raise ValueError(f"상품번호를 URL에서 찾을 수 없습니다: {url}")
    return match.group(1)


def _parse_incentive_date(date_str: str):
    """'26.6.1~6.7' → (date(2026,6,1), date(2026,6,7)), (None,None) on failure."""
    s = date_str.strip()
    if "~" not in s:
        return None, None
    left, right = s.split("~", 1)
    left_parts = left.strip().split(".")
    right_parts = right.strip().split(".")
    if len(left_parts) != 3:
        return None, None
    try:
        year = 2000 + int(left_parts[0])
        sm, sd = int(left_parts[1]), int(left_parts[2])
        if len(right_parts) == 2:
            em, ed = int(right_parts[0]), int(right_parts[1])
        elif len(right_parts) == 1:
            em, ed = sm, int(right_parts[0])
        else:
            return None, None
        ey = year + 1 if em < sm else year
        return date(year, sm, sd), date(ey, em, ed)
    except (ValueError, IndexError):
        return None, None


# ── 캠페인 목록 읽기 ─────────────────────────────────────
def load_campaigns() -> list:
    """
    구글 시트(MASTER_SHEET_URL)에서 캠페인 목록을 읽어옵니다.
    오늘이 시작일~종료일 사이인 캠페인만 반환합니다.

    시트 열 구조:
      A: No | B: 제목 | C: 시작일자 | D: 종료일자 | E: 링크 | F: 공유 구글스프레드
    """
    if not MASTER_SHEET_URL:
        print("[오류] .env 파일에 MASTER_SHEET_URL을 입력해주세요.")
        return []

    try:
        creds = Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=SCOPES)
        client = gspread.authorize(creds)
        sheet_id = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", MASTER_SHEET_URL).group(1)
        spreadsheet = client.open_by_key(sheet_id)
        gid_match = re.search(r"gid=(\d+)", MASTER_SHEET_URL)
        if gid_match:
            gid = int(gid_match.group(1))
            ws = next(s for s in spreadsheet.worksheets() if s.id == gid)
        else:
            ws = spreadsheet.sheet1
        rows = ws.get_all_records()
    except Exception as e:
        print(f"[오류] 캠페인 시트 읽기 실패: {e}")
        return []

    KST = timezone(timedelta(hours=9))
    today = datetime.now(KST).date()
    campaigns = []

    for row in rows:
        try:
            title     = str(row.get("제목") or "").strip()
            start_str = str(row.get("시작일자") or "").strip()
            end_str   = str(row.get("종료일자") or "").strip()
            url       = str(row.get("상품링크") or "").strip()
            sheet_url = str(row.get("데이터공유 구글스프레드_인플루언서전달링크") or "").strip()
            store     = str(row.get("스토어") or "").strip().lower()

            if not all([title, start_str, end_str, url, sheet_url, store]):
                continue

            # 재고 읽기 — 헤더에 "재고"가 포함된 열을 유연하게 탐색
            inventory_key = next((k for k in row.keys() if "재고" in k), None)
            inventory_raw = str(row.get(inventory_key) or "").strip() if inventory_key else ""
            if inventory_raw == "0" or inventory_raw.lower() in ("품절", "소진", "종료"):
                print(f"  [스킵] '{title}': 재고 소진으로 캠페인 제외")
                continue
            inventory = int(inventory_raw.replace(",", "")) if inventory_raw.replace(",", "").isdigit() else None

            if store not in STORE_CREDENTIALS:
                print(f"  [경고] 알 수 없는 스토어명: '{store}' (nutone/jdhealth/nutpet 중 하나여야 합니다)")
                continue

            api_id, api_secret = STORE_CREDENTIALS[store]
            if not api_id or not api_secret:
                print(f"  [경고] '{store}' API 키가 .env에 없습니다.")
                continue

            start_date = parse_date(start_str)
            end_date   = parse_date(end_str)

            if not (start_date <= today <= end_date):
                continue

            campaigns.append({
                "title":      title,
                "product_no": extract_product_no(url),
                "url":        url,
                "date_from":  start_date.strftime("%Y-%m-%d"),
                "date_to":    end_date.strftime("%Y-%m-%d"),
                "sheet_url":  sheet_url,
                "api_id":     api_id,
                "api_secret": api_secret,
                "inventory":  inventory,
            })
        except Exception as e:
            print(f"  [경고] 행 파싱 오류: {e}")

    return campaigns


# ── 파마브로스 파일공유 캠페인 읽기 ──────────────────────
def load_pharmabros_campaigns() -> list:
    """
    K열(파마브로스파일공유여부)이 '파마브로스파일공유'인 캠페인 반환.
    - 진행 중인 캠페인 (start <= today <= end)
    - 종료일 다음날 (end + 1일 == today) — 최종 업로드용

    반환 캠페인에 is_final 키 추가:
      is_final=True  → 종료 다음날 최종 업로드
      is_final=False → 진행 중 중간 업로드
    """
    if not MASTER_SHEET_URL:
        return []

    try:
        creds = Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=SCOPES)
        client = gspread.authorize(creds)
        sheet_id = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", MASTER_SHEET_URL).group(1)
        spreadsheet = client.open_by_key(sheet_id)
        gid_match = re.search(r"gid=(\d+)", MASTER_SHEET_URL)
        if gid_match:
            gid = int(gid_match.group(1))
            ws = next(s for s in spreadsheet.worksheets() if s.id == gid)
        else:
            ws = spreadsheet.sheet1
        rows = ws.get_all_records()
    except Exception as e:
        print(f"[오류] 파마브로스 캠페인 시트 읽기 실패: {e}")
        return []

    KST_zone = timezone(timedelta(hours=9))
    today    = datetime.now(KST_zone).date()
    campaigns = []

    for row in rows:
        try:
            # K열: 파마브로스파일공유여부
            pharmabros_flag = str(row.get("파마브로스파일공유여부") or "").strip()
            if not _is_pharmabros(pharmabros_flag):
                continue

            title     = str(row.get("제목") or "").strip()
            start_str = str(row.get("시작일자") or "").strip()
            end_str   = str(row.get("종료일자") or "").strip()
            url       = str(row.get("상품링크") or "").strip()
            store     = str(row.get("스토어") or "").strip().lower()

            if not all([title, start_str, end_str, url, store]):
                continue
            if store not in STORE_CREDENTIALS:
                continue

            api_id, api_secret = STORE_CREDENTIALS[store]
            if not api_id or not api_secret:
                continue

            start_date = parse_date(start_str)
            end_date   = parse_date(end_str)

            is_active    = start_date <= today <= end_date
            is_final_day = (today == end_date + timedelta(days=1))
            is_delete_day = (today == pharmabros.get_delete_date(end_str))

            if not (is_active or is_final_day or is_delete_day):
                continue

            campaigns.append({
                "title":         title,
                "product_no":    extract_product_no(url),
                "url":           url,
                "date_from":     start_date.strftime("%Y-%m-%d"),
                "date_to":       end_date.strftime("%Y-%m-%d"),
                "api_id":        api_id,
                "api_secret":    api_secret,
                "is_final":      is_final_day,
                "is_delete_day": is_delete_day,
            })
        except Exception as e:
            print(f"  [경고] 파마브로스 행 파싱 오류: {e}")

    return campaigns


# ── 캠페인 전체 목록 읽기 (종료된 것 포함) ───────────────
def load_all_campaigns() -> list:
    """시작일이 오늘 이전인 모든 캠페인 반환 (종료 포함)."""
    if not MASTER_SHEET_URL:
        return []

    try:
        creds = Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=SCOPES)
        client = gspread.authorize(creds)
        sheet_id = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", MASTER_SHEET_URL).group(1)
        spreadsheet = client.open_by_key(sheet_id)
        gid_match = re.search(r"gid=(\d+)", MASTER_SHEET_URL)
        if gid_match:
            gid = int(gid_match.group(1))
            ws = next(s for s in spreadsheet.worksheets() if s.id == gid)
        else:
            ws = spreadsheet.sheet1
        rows = ws.get_all_records()
    except Exception as e:
        print(f"[오류] 캠페인 시트 읽기 실패: {e}")
        return []

    KST = timezone(timedelta(hours=9))
    today = datetime.now(KST).date()
    campaigns = []

    for row in rows:
        try:
            title     = str(row.get("제목") or "").strip()
            start_str = str(row.get("시작일자") or "").strip()
            end_str   = str(row.get("종료일자") or "").strip()
            url       = str(row.get("상품링크") or "").strip()
            store     = str(row.get("스토어") or "").strip().lower()

            if not all([title, start_str, end_str, url, store]):
                continue
            if store not in STORE_CREDENTIALS:
                continue

            api_id, api_secret = STORE_CREDENTIALS[store]
            if not api_id or not api_secret:
                continue

            start_date = parse_date(start_str)
            end_date   = parse_date(end_str)

            if start_date > today:
                continue

            sheet_url = str(row.get("데이터공유 구글스프레드_인플루언서전달링크") or "").strip()
            campaigns.append({
                "title":      title,
                "product_no": extract_product_no(url),
                "url":        url,
                "sheet_url":  sheet_url,
                "date_from":  start_date.strftime("%Y-%m-%d"),
                "date_to":    end_date.strftime("%Y-%m-%d"),
                "api_id":     api_id,
                "api_secret": api_secret,
                "store":      store,
                "is_ended":   end_date < today,
            })
        except Exception as e:
            print(f"  [경고] 행 파싱 오류: {e}")

    return campaigns


# ── 인센티브 시트에서 캠페인 읽기 ─────────────────────────
def load_incentive_campaigns(min_year_month: str = "2026-05") -> list:
    """인센티브정산 시트(COMM_SHEET_URL)에서 공구 캠페인 목록 읽기.
    2행 1세트(공구+공구링크) 형식을 파싱한다.
    """
    if not COMM_SHEET_URL:
        return []
    try:
        creds = Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=SCOPES)
        client = gspread.authorize(creds)
        sid = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", COMM_SHEET_URL).group(1)
        ss = client.open_by_key(sid)
    except Exception as e:
        print(f"  [인센티브시트 읽기 실패] {e}")
        return []

    KST = timezone(timedelta(hours=9))
    today = datetime.now(KST).date()
    raw = []

    for ws_tab in ss.worksheets():
        ym = _tab_year_month(ws_tab.title)
        if not ym or ym < min_year_month:
            continue
        try:
            rows = ws_tab.get_all_values()
        except Exception:
            continue

        i = 0
        while i < len(rows):
            row = rows[i]
            a = row[0].strip() if len(row) > 0 else ""
            b = row[1].strip() if len(row) > 1 else ""
            d = row[3].strip() if len(row) > 3 else ""

            if b == "공구" and "~" in a:
                start_d, end_d = _parse_incentive_date(a)
                channel = d
                j = i + 1
                while j < len(rows):
                    nr = rows[j]
                    nb = nr[1].strip() if len(nr) > 1 else ""
                    na = nr[0].strip() if len(nr) > 0 else ""
                    if nb == "공구" or (na and "~" in na):
                        break
                    if nb == "공구링크":
                        pname = nr[4].strip() if len(nr) > 4 else ""
                        purl  = nr[5].strip() if len(nr) > 5 else ""
                        # M열(index 12) 수수료 추출
                        comm_raw = nr[12].strip() if len(nr) > 12 else ""
                        comm_m = re.search(r'(\d+(?:\.\d+)?)', comm_raw)
                        comm_val = float(comm_m.group(1)) if comm_m else ""
                        if start_d and end_d and purl:
                            sm = re.search(r'brand\.naver\.com/(\w+)', purl)
                            store = sm.group(1) if sm else ""
                            pm = re.search(r'/products/(\d+)', purl)
                            pno = pm.group(1) if pm else ""
                            if store in STORE_CREDENTIALS and pno:
                                aid, asec = STORE_CREDENTIALS[store]
                                if aid and asec:
                                    raw.append({
                                        "channel_name": channel,
                                        "product_name": pname,
                                        "incentive_comm": comm_val,
                                        "product_no":   pno,
                                        "url":          purl.split("?")[0],
                                        "date_from":    start_d.strftime("%Y-%m-%d"),
                                        "date_to":      end_d.strftime("%Y-%m-%d"),
                                        "api_id":       aid,
                                        "api_secret":   asec,
                                        "store":        store,
                                        "sheet_url":    "",
                                        "is_ended":     end_d < today,
                                    })
                    j += 1
                i = j
            else:
                i += 1

    from collections import Counter
    ch_counts = Counter(c["channel_name"] for c in raw)
    for c in raw:
        if ch_counts[c["channel_name"]] > 1:
            c["title"] = f"{c['channel_name']},{c['product_name']}"
        else:
            c["title"] = c["channel_name"]

    return raw


def _calc_totals(sales_data: list) -> tuple:
    """판매 데이터에서 총 주문수, 총 제품수 계산 (개별 시트와 동일 방식)."""
    aggregated = sheets._aggregate(sales_data)
    total_orders = sum(r["daily_orders"] for r in aggregated)
    total_products = sum(r["daily_products"] for r in aggregated)
    return total_orders, total_products


def _read_master_sheet1_lm() -> dict:
    """시트1의 I열(매출), J열(공구수수료) 읽기.
    반환: {제목: {"revenue": int|"", "commission": float|""}}
    """
    try:
        creds = Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=SCOPES)
        client = gspread.authorize(creds)
        sheet_id = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", MASTER_SHEET_URL).group(1)
        ss = client.open_by_key(sheet_id)
        ws = ss.sheet1
        rows = ws.get_all_records()
        result = {}
        for row in rows:
            title = str(row.get("제목", "")).strip()
            if not title:
                continue
            # I열 매출 (숫자, 쉼표 제거)
            raw_rev = str(row.get("매출", "")).replace(",", "").strip()
            revenue = int(raw_rev) if raw_rev.lstrip("-").isdigit() else ""
            # J열 공구수수료 — 숫자 추출 ("40", "40%", "공구수수료40%" 모두 처리)
            raw_comm = str(row.get("공구수수료(%,vat포함)", "")).strip()
            m = re.search(r"(\d+(?:\.\d+)?)", raw_comm)
            commission = float(m.group(1)) if m else ""
            result[title] = {"revenue": revenue, "commission": commission}
        return result
    except Exception as e:
        print(f"  [시트1 읽기 실패] {e}")
        return {}


def _extract_commission_raw(raw: str):
    """수수료 문자열에서 숫자 추출. '공구수수료40%', '40%', '40' 모두 처리."""
    s = str(raw).strip()
    m = re.search(r'공구수수료\s*(\d+(?:\.\d+)?)\s*%', s)
    if m:
        return float(m.group(1))
    if '컨텐츠' not in s:
        m = re.search(r'(\d+(?:\.\d+)?)', s)
        if m:
            return float(m.group(1))
    return ""


def _tab_year_month(tab_title: str) -> str:
    """탭 제목에서 'YYYY-MM' 추출. '26.4', '4월' 등 처리."""
    t = tab_title.strip()
    m = re.match(r'(\d{2})\.(\d{1,2})', t)
    if m:
        return f"20{m.group(1)}-{int(m.group(2)):02d}"
    m2 = re.match(r'(\d{1,2})\s*월', t)
    if m2:
        return f"2026-{int(m2.group(1)):02d}"
    return ""


def _read_comm_map() -> tuple:
    """공구수수료 시트에서 채널명→수수료 매핑 읽기.
    반환: (comm_map, comm_by_product_no)
      comm_map             = {채널명(소문자): 수수료율(float)}
      comm_by_product_no   = {상품번호: [(수수료율, "YYYY-MM"), ...]}
    """
    comm_map = {}
    comm_by_product_no = {}
    if not COMM_SHEET_URL:
        return comm_map, comm_by_product_no
    try:
        creds  = Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=SCOPES)
        client = gspread.authorize(creds)
        comm_sid = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", COMM_SHEET_URL).group(1)
        ss_comm  = client.open_by_key(comm_sid)

        def _pno(url):
            m = re.search(r"/products/(\d+)", str(url))
            return m.group(1) if m else ""

        for ws_tab in ss_comm.worksheets():
            title_lower = ws_tab.title.lower()
            if not any(x in title_lower for x in ["월", ".1", ".2", ".3", ".4", ".5", ".6", ".7", ".8", ".9"]):
                continue
            ym = _tab_year_month(ws_tab.title)
            try:
                tab_rows = ws_tab.get_all_values()
            except Exception:
                continue
            i = 0
            while i < len(tab_rows):
                row = tab_rows[i]
                b_col = row[1].strip() if len(row) > 1 else ""
                d_col = row[3].strip() if len(row) > 3 else ""
                if b_col == "공구" and d_col:
                    channel = d_col
                    j = i + 1
                    while j < len(tab_rows):
                        nrow = tab_rows[j]
                        nb = nrow[1].strip() if len(nrow) > 1 else ""
                        if nb == "공구":
                            break
                        if nb == "공구링크":
                            m_val = nrow[12].strip() if len(nrow) > 12 else ""
                            comm  = _extract_commission_raw(m_val)
                            if comm != "":
                                if channel:
                                    comm_map[channel.lower()] = comm
                                for cell in nrow:
                                    pno = _pno(str(cell))
                                    if pno and ym:
                                        comm_by_product_no.setdefault(pno, []).append((comm, ym))
                        j += 1
                    i = j
                else:
                    i += 1
    except Exception as e:
        print(f"  [공구수수료 시트 읽기 실패] {e}")
    return comm_map, comm_by_product_no


def update_summary_tab():
    """마스터 시트의 '캠페인 실적' 탭 업데이트."""
    KST = timezone(timedelta(hours=9))
    all_campaigns = load_all_campaigns()

    # 인센티브 시트에서 추가 캠페인 병합 (시트1에 없는 것만)
    incentive = load_incentive_campaigns()
    if incentive:
        seen = {(c["product_no"], c["date_from"], c["date_to"]) for c in all_campaigns}
        # 시트1 sheet_url 매핑 (인센티브 캠페인에 연결용)
        s1_sheet_by_pno = {}
        for c in all_campaigns:
            if c.get("sheet_url"):
                s1_sheet_by_pno.setdefault(c["product_no"], []).append(c)
        added = 0
        for ic in incentive:
            key = (ic["product_no"], ic["date_from"], ic["date_to"])
            if key not in seen:
                # 시트1에서 같은 상품번호의 sheet_url 연결
                for s1c in s1_sheet_by_pno.get(ic["product_no"], []):
                    if s1c["date_from"] == ic["date_from"] or s1c["date_to"] == ic["date_to"]:
                        ic["sheet_url"] = s1c["sheet_url"]
                        break
                all_campaigns.append(ic)
                seen.add(key)
                added += 1
        if added:
            print(f"  [인센티브 시트] {added}개 캠페인 추가 로드")

    if not all_campaigns:
        return

    existing = sheets.read_summary_tab(MASTER_SHEET_URL)
    today_str = datetime.now(KST).date().strftime("%Y-%m-%d")
    changed = False

    # 1) 기존 행: 상태 빈칸 채우기 + URL 정리
    for r in existing.values():
        if not r.get("status"):
            r["status"] = "완료" if str(r.get("date_to", "")) < today_str else "진행중"
            changed = True
        if r.get("url") and "?" in r["url"]:
            r["url"] = r["url"].split("?")[0]
            changed = True

    # 2) 새로 처리할 캠페인:
    #    - 종료됐고 탭에 없는 신규
    #    - 제품명이 없거나 캠페인 제목과 동일한 경우 (잘못된 짧은 이름)
    fetch_campaigns = [
        c for c in all_campaigns
        if (c["is_ended"] and c["title"] not in existing)
        or (c["title"] in existing
            and (not existing[c["title"]].get("product_name")
                 or existing[c["title"]].get("product_name") == c["title"]))
    ]

    new_rows = []
    if fetch_campaigns:
        print(f"\n  [실적 집계] 캠페인 {len(fetch_campaigns)}개 처리 중...")
        for campaign in fetch_campaigns:
            clean_url = campaign["url"].split("?")[0]
            clean_url = campaign["url"].split("?")[0]
            # 주문 상세에서 product_name 추출
            try:
                sales, product_name = naver_api.get_sales_data(
                    client_id=campaign["api_id"],
                    client_secret=campaign["api_secret"],
                    product_no=campaign["product_no"],
                    date_from=campaign["date_from"],
                    date_to=campaign["date_to"],
                )
            except Exception:
                sales, product_name = [], ""
            # 주문 0건이면 상품 API로 폴백
            if not product_name:
                product_name = naver_api.get_product_name(
                    campaign["api_id"], campaign["api_secret"], campaign["product_no"]
                )
            # API에서도 못 가져오면 인센티브 시트 제품명 사용
            if not product_name:
                product_name = campaign.get("product_name", "")
            # 합계는 개별 시트에서 읽기 (정확도 우선)
            try:
                total_orders, total_products, _ = sheets.read_totals_from_sheet(campaign["sheet_url"])
            except Exception:
                total_orders, total_products = _calc_totals(sales)

            new_rows.append({
                "title":          campaign["title"],
                "product_name":   product_name,
                "url":            clean_url,
                "store":          campaign["store"],
                "date_from":      campaign["date_from"],
                "date_to":        campaign["date_to"],
                "total_orders":   total_orders,
                "total_products": total_products,
                "status":         "완료" if campaign["is_ended"] else "진행중",
                "updated_at":     datetime.now(KST).strftime("%Y-%m-%d %H:%M"),
            })
        changed = True

    if not changed and not new_rows:
        # 매출은 있는데 수수료가 없는 기존 행이 있으면 채우기 위해 계속 진행
        needs_fill = any(
            isinstance(row.get("revenue"), int) and row["revenue"] > 0
            and row.get("commission", "") == ""
            for row in existing.values()
        )
        if not needs_fill:
            return

    new_titles = {r["title"] for r in new_rows}
    kept_rows = [r for r in existing.values() if r["title"] not in new_titles]
    all_rows = new_rows + kept_rows
    all_rows.sort(key=lambda r: str(r.get("date_to", "")), reverse=True)

    # ── (product_no, date_from) 기준 중복 제거 ───────────────────
    # 같은 캠페인이 sheet1/인센티브 시트에서 제목·종료일이 달라도 하나만 유지
    _pno_seen: dict = {}
    _deduped: list = []
    for r in all_rows:
        _m = re.search(r"/products/(\d+)", str(r.get("url", "")))
        _pno = _m.group(1) if _m else ""
        _pkey = (_pno, str(r.get("date_from", ""))) if _pno else ("__title__", r.get("title", ""))
        if _pkey not in _pno_seen:
            _pno_seen[_pkey] = True
            _deduped.append(r)
    all_rows = _deduped

    # ── L/M/N/O열 계산 ────────────────────────────────────────
    # 이익계산참고사항 로드 (시트 읽기, 1회)
    profit_params = []
    try:
        profit_params = sheets.read_profit_params(MASTER_SHEET_URL)
    except Exception as e:
        print(f"  [이익참고] 읽기 실패: {e}")

    # 시트1 J열(공구수수료) 로드 (시트 읽기, 1회)
    master1_lm = _read_master_sheet1_lm()

    # 공구수수료 별도 시트 로드 (시트1에 수수료 없는 행 대비 fallback)
    comm_map, comm_by_product_no = _read_comm_map()

    # 새로 집계되는 캠페인 제목 집합 (이 타이밍에만 매출 API 호출)
    fetch_titles   = {r["title"] for r in new_rows}
    title_to_campaign = {c["title"]: c for c in all_campaigns}

    row_extras = {}
    for row in all_rows:
        title        = row["title"]
        product_name = row.get("product_name", "") or title
        pp           = sheets._match_profit(product_name, profit_params)

        # ── 매출(L): 이미 계산된 캠페인은 캐시 그대로 ──────────
        revenue_refetched = False
        if title not in fetch_titles:
            # kept_rows: 기존 L열 값 재사용 (API 호출 없음)
            revenue = row.get("revenue", "")
            # 매출 누락 + 종료된 캠페인 → API 재조회
            if revenue == "" and str(row.get("date_to", "")) < today_str:
                campaign = title_to_campaign.get(title, {})
                if campaign and campaign.get("api_id"):
                    try:
                        revenue = naver_api.get_campaign_revenue(
                            campaign["api_id"], campaign["api_secret"],
                            campaign["product_no"],
                            campaign.get("date_from", str(row.get("date_from", ""))),
                            campaign.get("date_to", str(row.get("date_to", ""))),
                        )
                        if isinstance(revenue, int):
                            print(f"  [매출 재조회] {title}: {revenue:,}원")
                            revenue_refetched = True
                    except Exception as e:
                        print(f"  [매출 재조회 실패] {title}: {e}")
        else:
            # fetch_campaigns: 종료 시 1회만 API 조회
            s1_rev = master1_lm.get(title, {}).get("revenue", "")
            if isinstance(s1_rev, int) and s1_rev > 0:
                revenue = s1_rev          # 시트1에 수기 입력값 우선
            else:
                campaign = title_to_campaign.get(title, {})
                if campaign:
                    try:
                        revenue = naver_api.get_campaign_revenue(
                            campaign["api_id"], campaign["api_secret"],
                            campaign["product_no"],
                            campaign["date_from"], campaign["date_to"],
                        )
                    except Exception as e:
                        print(f"  [매출조회 실패] {title}: {e}")
                        revenue = ""
                else:
                    revenue = ""

        # ── 공구수수료(M): 시트1 J열 우선, 없으면 캐시, 그래도 없으면 공구수수료 시트 fallback ──
        s1_comm    = master1_lm.get(title, {}).get("commission", "")
        commission = s1_comm if s1_comm != "" else row.get("commission", "")

        if commission == "" and comm_map:
            # 1순위: 채널명 정확 매칭
            commission = comm_map.get(title.lower(), "")
            # 2순위: 채널명 부분 매칭
            if commission == "":
                t_lower = title.lower()
                for key, val in comm_map.items():
                    if (len(t_lower) >= 3 and t_lower in key) or (len(key) >= 3 and key in t_lower):
                        commission = val
                        print(f"  [수수료 부분매칭] '{title}' ↔ '{key}' → {val}%")
                        break
            # 3순위: 상품번호 + 캠페인 월 매칭
            if commission == "":
                row_url = row.get("url", "") or ""
                pno_m = re.search(r"/products/(\d+)", str(row_url))
                if pno_m:
                    pno        = pno_m.group(1)
                    campaign_ym = str(row.get("date_from", ""))[:7]
                    for c, ym in comm_by_product_no.get(pno, []):
                        if ym == campaign_ym:
                            commission = c
                            print(f"  [수수료 상품번호+월 매칭] '{title}' → {c}%")
                            break

        # 4순위: 인센티브 시트에서 직접 가져온 수수료
        if commission == "":
            campaign_info_comm = title_to_campaign.get(title, {})
            ic_comm = campaign_info_comm.get("incentive_comm", "")
            if ic_comm != "":
                commission = ic_comm
                print(f"  [수수료 인센티브시트] '{title}' → {ic_comm}%")

        # ── 옵션별 원가 계산 (멀티제품 캠페인) ──────────────────
        # 새로 종료된 캠페인(fetch_titles)에만 개별 시트 읽기 실행
        # kept_rows는 이미 N 수식이 있으므로 매 시간 재조회 불필요
        option_cost_total = 0
        option_delivery   = 0
        option_ch_comm    = 0.0
        opt_matched_names = []
        campaign_info = title_to_campaign.get(title, {})
        c_sheet_url   = campaign_info.get("sheet_url", "")
        if (title in fetch_titles or revenue_refetched) and c_sheet_url and isinstance(revenue, int) and revenue > 0:
            try:
                option_totals = sheets.read_option_totals_from_sheet(c_sheet_url)
                if len(option_totals) >= 2:
                    opt_res = sheets.calc_option_cost(option_totals, profit_params)
                    if opt_res["total_cost"] > 0:
                        option_cost_total = opt_res["total_cost"]
                        option_delivery   = opt_res["delivery"]
                        option_ch_comm    = opt_res["ch_comm"]
                        opt_matched_names = opt_res["matched_names"]
                        print(f"  [멀티원가] {title[:30]}: {' + '.join(opt_matched_names)} = {option_cost_total:,}원")
            except Exception as e:
                print(f"  [옵션별 원가 실패] {title}: {e}")

        # P열 매칭 제품명 (멀티제품이면 "+", 단일이면 단독)
        if opt_matched_names:
            matched_name = " + ".join(opt_matched_names)
        else:
            matched_name = str(pp.get("name", "")) if pp else ""

        row_extras[title] = {
            "revenue":           revenue,
            "commission":        commission,
            "profit_params":     pp,
            "matched_name":      matched_name,
            "option_cost_total": option_cost_total,
            "option_delivery":   option_delivery,
            "option_ch_comm":    option_ch_comm,
        }

    sheets.write_summary_tab(MASTER_SHEET_URL, all_rows, row_extras=row_extras)
    print(f"  → 캠페인 실적 탭 업데이트 완료\n")


# ── 메인 실행 ────────────────────────────────────────────
def run_once():
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"\n{'='*55}")
    print(f"  실행 시작: {now_str}")
    print(f"{'='*55}")

    if not MASTER_SHEET_URL:
        print("[오류] .env 파일에 MASTER_SHEET_URL을 입력해주세요.")
        return

    campaigns = load_campaigns()

    if not campaigns:
        print(f"  오늘({date.today()}) 진행 중인 캠페인이 없습니다.")
        return

    print(f"  진행 중인 캠페인: {len(campaigns)}개\n")

    soldout_notified = _load_soldout_notified()

    for i, campaign in enumerate(campaigns, 1):
        print(f"[{i}/{len(campaigns)}] {campaign['title'][:45]}")
        try:
            sales, product_name = naver_api.get_sales_data(
                client_id=campaign["api_id"],
                client_secret=campaign["api_secret"],
                product_no=campaign["product_no"],
                date_from=campaign["date_from"],
                date_to=campaign["date_to"],
            )
            remaining = sheets.write_to_sheet(
                spreadsheet_url=campaign["sheet_url"],
                product_title=campaign["title"],
                sales_data=sales,
                date_from=campaign["date_from"],
                date_to=campaign["date_to"],
                product_url=campaign.get("url", ""),
                product_name=product_name,
                inventory=campaign.get("inventory"),
            )
            product_no = campaign["product_no"]
            if remaining is not None:
                if remaining <= 0:
                    # 남은재고 0 → 품절 알림 (아직 알림 안 보낸 경우만)
                    if product_no not in soldout_notified:
                        inv = campaign.get("inventory") or 0
                        send_telegram(
                            f"🔴 [품절 알림]\n\n"
                            f"📦 상품명: {campaign['title'][:40]}\n"
                            f"📊 초기재고: {inv:,}개\n"
                            f"✅ 판매 완료: 재고 소진\n"
                            f"🏪 스토어: {next((k for k, v in STORE_CREDENTIALS.items() if v[0] == campaign.get('api_id')), '')}\n\n"
                            f"🕐 {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                        )
                        soldout_notified.add(product_no)
                        _save_soldout_notified(soldout_notified)
                        print(f"  [품절 알림] 텔레그램 발송 완료")
                else:
                    # 재고 보충됨 → 알림 기록 초기화 (다음 품절 시 재알림 가능)
                    if product_no in soldout_notified:
                        soldout_notified.discard(product_no)
                        _save_soldout_notified(soldout_notified)
                        print(f"  [재고 보충 감지] 품절 알림 기록 초기화")
            print(f"  완료\n")
        except Exception as e:
            print(f"  [오류] {e}\n")
            err_str = str(e)
            err_lower = err_str.lower()
            # 품절·판매중지·404 오류는 Telegram 알림 없이 조용히 넘어감
            if ("404" in err_str or "not found" in err_lower
                    or "품절" in err_str or "sold_out" in err_lower
                    or "no such product" in err_lower):
                print(f"  [스킵] 품절 또는 판매중지 상품으로 판단 — 오류 알림 생략\n")
                continue
            if "403" in err_str or "Forbidden" in err_str:
                cause = "네이버 API 인증 오류 (IP 차단 또는 API 키 만료)"
                action = "→ 네이버 커머스 API에서 IP 및 API 키를 확인해주세요"
            elif "401" in err_str or "Unauthorized" in err_str:
                cause = "네이버 API 키 오류 (ID 또는 Secret 불일치)"
                action = "→ .env 파일의 API 키를 확인해주세요"
            elif "spreadsheet" in err_str.lower() or "gspread" in err_str.lower():
                cause = "구글 시트 접근 오류 (권한 또는 URL 문제)"
                action = "→ 인플루언서 구글 시트가 서비스 계정과 공유되어 있는지 확인해주세요"
            elif "products" in err_str or "상품번호" in err_str:
                cause = "상품 URL에서 상품번호를 찾을 수 없음"
                action = "→ 캠페인 시트의 상품링크 URL을 확인해주세요"
            else:
                cause = f"알 수 없는 오류: {err_str[:100]}"
                action = "→ 개발자에게 문의해주세요"

            send_telegram(
                f"⚠️ [인플루언서 프로그램 오류]\n\n"
                f"📦 상품명: {campaign['title'][:40]}\n"
                f"🔗 인플루언서 시트: {campaign.get('sheet_url', '')}\n"
                f"🏪 스토어: {next((k for k, v in STORE_CREDENTIALS.items() if v[0] == campaign.get('api_id')), campaign.get('api_id', ''))}\n\n"
                f"❌ 원인: {cause}\n"
                f"{action}\n\n"
                f"🕐 발생시각: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            )

    update_summary_tab()

    # ── 파마브로스 파일공유 ───────────────────────────────
    _run_pharmabros_if_needed()

    # ── 파마브로스 정산 백필 (종료 캠페인 중 미처리건) ────
    _backfill_pharmabros_settlements()

    # ── 사은품 추첨 (종료 +1~7일 캠페인, L열 있을 때) ────
    _run_saeunpum_lottery()

    print(f"{'='*55}\n")


# ── 파마브로스 정산 계산 ──────────────────────────────────
def _save_pharmabros_option_prices(campaign) -> dict:
    """캠페인 중간 시점에 옵션가격 JSON 저장. 이미 저장됐으면 기존값 반환."""
    PHARMABROS_PRICES_DIR.mkdir(exist_ok=True)
    price_file = PHARMABROS_PRICES_DIR / f"{campaign['product_no']}.json"

    if price_file.exists():
        return json.loads(price_file.read_text(encoding="utf-8"))

    KST = timezone(timedelta(hours=9))
    start = datetime.strptime(campaign["date_from"], "%Y-%m-%d").date()
    end = datetime.strptime(campaign["date_to"], "%Y-%m-%d").date()
    today = datetime.now(KST).date()
    mid = start + (end - start) // 2

    if today < mid:
        return {}

    prices = naver_api.get_product_option_prices(
        campaign["api_id"], campaign["api_secret"], campaign["product_no"]
    )
    if prices:
        price_file.write_text(json.dumps(prices, ensure_ascii=False), encoding="utf-8")
        print(f"  [옵션가격 저장] {campaign['title'][:30]}: {len(prices)}개 옵션")
    return prices


def _match_option_price(opt_name: str, prices: dict) -> int:
    """옵션명에서 BOX 수량 추출 후 가격 매칭."""
    m = re.search(r'(\d+)\s*(?:bx|box|박스)', opt_name, re.IGNORECASE)
    if not m:
        if len(prices) == 1:
            return list(prices.values())[0]
        return 0
    opt_num = int(m.group(1))
    for price_name, price in prices.items():
        m2 = re.search(r'(\d+)\s*(?:bx|box|박스)', price_name, re.IGNORECASE)
        if m2 and int(m2.group(1)) == opt_num:
            return price
    return 0


def _auto_save_pharmabros_prices(campaign):
    """주문 API에서 옵션별 단가를 자동 추출해 JSON 저장 (가격 파일 없을 때만)."""
    price_file = PHARMABROS_PRICES_DIR / f"{campaign['product_no']}.json"
    if price_file.exists():
        return
    try:
        orders = naver_api.get_pharmabros_orders(
            campaign["api_id"], campaign["api_secret"],
            campaign["product_no"],
            campaign["date_from"], campaign["date_to"],
        )
    except Exception as e:
        print(f"  [자동가격저장] 주문 조회 실패: {e}")
        return

    from collections import Counter
    opt_prices: dict = {}
    for o in orders:
        status = str(o.get("주문상태", ""))
        if "취소" in status or "CANCEL" in status.upper():
            continue
        opt = str(o.get("옵션", "")).strip()
        amt = int(o.get("단가", 0) or 0)
        if opt and amt > 0:
            opt_prices.setdefault(opt, []).append(amt)

    if not opt_prices:
        print(f"  [자동가격저장] {campaign['title'][:20]}: 유효 주문 없음")
        return

    prices = {}
    for opt, vals in opt_prices.items():
        unique = set(vals)
        if len(unique) > 1:
            print(f"  ⚠️ [단가불일치] '{opt}': {sorted(unique)} → 최댓값({max(vals)}) 선택")
        prices[opt] = max(vals)  # 불일치 시 최댓값(정가에 가까운 값) 강제 선택
    price_file.parent.mkdir(exist_ok=True)
    price_file.write_text(json.dumps(prices, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  [자동가격저장] {campaign['title'][:20]}: {prices}")


def _calc_pharmabros_settlement(campaign) -> tuple:
    """Drive xlsx에서 주문 데이터를 읽어 정산금액 계산.
    고정 옵션가격이 있으면 적용. 반환: (총액, xlsx행리스트)
    """
    _auto_save_pharmabros_prices(campaign)   # 가격 파일 없으면 자동 생성
    rows = pharmabros.read_xlsx_rows_from_drive(
        client_id=PHARMABROS_OAUTH_CLIENT_ID,
        client_secret=PHARMABROS_OAUTH_CLIENT_SECRET,
        refresh_token=PHARMABROS_OAUTH_REFRESH_TOKEN,
        folder_id=PHARMABROS_DRIVE_FOLDER_ID,
        title=campaign["title"],
    )
    if not rows:
        print(f"  [파마브로스정산] Drive 파일 없음: {campaign['title'][:30]}")
        return 0, []

    # 고정 옵션가격 로드
    price_file = PHARMABROS_PRICES_DIR / f"{campaign['product_no']}.json"
    fixed_prices = {}
    if price_file.exists():
        fixed_prices = json.loads(price_file.read_text(encoding="utf-8"))

    total = 0
    for r in rows:
        status = str(r.get("주문상태", ""))
        if "취소" in status or "CANCEL" in status.upper():
            continue

        qty = int(r.get("주문수량", 0) or 0)
        # API 기반 정가(fixed_prices)를 항상 우선 사용, 없을 때만 xlsx 단가 참조
        unit_price = 0
        if fixed_prices:
            unit_price = _match_option_price(str(r.get("옵션", "")), fixed_prices)
        if not unit_price:
            raw_price = r.get("단가", 0) or 0
            try:
                unit_price = int(raw_price)
            except (ValueError, TypeError):
                unit_price = 0
        if unit_price:
            r["단가"] = unit_price

        total += unit_price * qty

    print(f"  [파마브로스정산] {campaign['title'][:30]}: {len(rows)}건, 합계 {total:,}원")
    return total, rows


def _write_pharmabros_revenue_to_i(campaign):
    """시트1 I열에 네이버 API 매출(취소/반품 제외) 자동 기재."""
    try:
        revenue = naver_api.get_campaign_revenue(
            campaign["api_id"], campaign["api_secret"],
            campaign["product_no"], campaign["date_from"], campaign["date_to"],
        )
        if not revenue:
            return
        creds = Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=SCOPES)
        client = gspread.authorize(creds)
        sheet_id = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", MASTER_SHEET_URL).group(1)
        ws = client.open_by_key(sheet_id).sheet1
        rows = ws.get_all_values()
        for i, row in enumerate(rows):
            if len(row) > 1 and row[1].strip() == campaign["title"]:
                existing_val = str(row[8]).replace(",", "").strip() if len(row) > 8 else ""
                if existing_val and existing_val != "0":
                    return
                ws.update_cell(i + 1, 9, revenue)
                print(f"  [파마브로스] I열 매출 기재: {campaign['title'][:30]} → {revenue:,}원")
                return
    except Exception as e:
        print(f"  [파마브로스] I열 기재 실패: {e}")


def _write_pharmabros_settlement(title: str, settlement: int, xlsx_rows: list):
    """파마브로스정산 탭에 Drive xlsx + 고정가격 기록."""
    try:
        creds = Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=SCOPES)
        client = gspread.authorize(creds)
        sheet_id = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", MASTER_SHEET_URL).group(1)
        ss = client.open_by_key(sheet_id)

        HEADER = ["제목", "주문번호", "주문일시", "주문상태", "옵션", "주문수량", "단가", "단가총합"]
        end_col = chr(64 + len(HEADER))

        try:
            pb_ws = ss.worksheet("파마브로스정산")
        except gspread.WorksheetNotFound:
            pb_ws = ss.add_worksheet(title="파마브로스정산", rows=500, cols=len(HEADER))

        existing = pb_ws.get_all_values()
        new_rows = [HEADER]
        for r in existing[1:]:
            if r and str(r[0]).strip() != title:
                new_rows.append(r)

        for row in xlsx_rows:
            try:
                unit_price = int(row.get("단가", 0) or 0)
            except (ValueError, TypeError):
                unit_price = 0
            try:
                qty = int(row.get("주문수량", 0) or 0)
            except (ValueError, TypeError):
                qty = 0
            new_rows.append([
                title,
                row.get("주문번호", ""),
                row.get("주문일시", ""),
                row.get("주문상태", ""),
                row.get("옵션", ""),
                qty,
                unit_price,
                unit_price * qty,
            ])

        pb_ws.clear()
        pb_ws.update(f"A1:{end_col}{len(new_rows)}", new_rows)
        print(f"  [파마브로스정산] {title[:30]}: {len(xlsx_rows)}건 기록, 합계 {settlement:,}원")

    except Exception as e:
        print(f"  [파마브로스정산] 기재 실패: {e}")


def _backfill_pharmabros_settlements():
    """종료된 파마브로스 캠페인 중 정산 탭에 데이터 없는 건 일괄 계산."""
    if not MASTER_SHEET_URL:
        return
    try:
        creds = Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=SCOPES)
        client = gspread.authorize(creds)
        sheet_id = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", MASTER_SHEET_URL).group(1)
        ss = client.open_by_key(sheet_id)
        ws = ss.sheet1
        rows = ws.get_all_records()
    except Exception as e:
        print(f"  [파마브로스 백필] 시트 읽기 실패: {e}")
        return

    # (기존 데이터가 있어도 항상 재계산 — _write_pharmabros_settlement에서 덮어씀)

    KST = timezone(timedelta(hours=9))
    today = datetime.now(KST).date()
    processed = 0

    for row in rows:
        k_val = re.sub(r"\s+", "", str(row.get("파마브로스파일공유여부", "")))
        if k_val != "파마브로스파일공유":
            continue

        title = str(row.get("제목", "")).strip()
        if not title:
            continue

        end_str = str(row.get("종료일자", "")).strip()
        if not end_str:
            continue
        try:
            end_date = parse_date(end_str)
        except ValueError:
            continue
        if end_date >= today:
            continue

        url   = str(row.get("상품링크", "")).strip()
        store = str(row.get("스토어", "")).strip().lower()
        start_str = str(row.get("시작일자", "")).strip()

        if store not in STORE_CREDENTIALS or not url:
            continue
        api_id, api_secret = STORE_CREDENTIALS[store]
        if not api_id or not api_secret:
            continue

        try:
            product_no = extract_product_no(url)
        except ValueError:
            continue

        campaign = {
            "title": title,
            "product_no": product_no,
            "url": url,
            "date_from": parse_date(start_str).strftime("%Y-%m-%d"),
            "date_to": end_date.strftime("%Y-%m-%d"),
            "api_id": api_id,
            "api_secret": api_secret,
            "store": store,
        }

        print(f"  [파마브로스 백필] {title[:30]} 처리 중...")
        _write_pharmabros_revenue_to_i(campaign)
        settlement, breakdown = _calc_pharmabros_settlement(campaign)
        if settlement > 0:
            _write_pharmabros_settlement(title, settlement, breakdown)
            processed += 1

    if processed:
        print(f"  [파마브로스 백필] {processed}개 캠페인 정산 완료")


def _upload_pharmabros_settlement_pdf(campaign, total_payment: int):
    """파마브로스 정산서 HTML 생성 후 Drive '정산서' 폴더에 자동 업로드.
    total_payment = 파마브로스 고정 옵션가 기준 합계금액 (취소 제외).
    정산금액은 내부에서 J열 수수료율 적용해 계산.
    """
    if not PHARMABROS_SETTLEMENT_FOLDER_ID:
        return
    if not all([PHARMABROS_OAUTH_CLIENT_ID, PHARMABROS_OAUTH_CLIENT_SECRET, PHARMABROS_OAUTH_REFRESH_TOKEN]):
        return
    if not total_payment:
        return

    title     = campaign["title"]
    date_from = campaign["date_from"]
    date_to   = campaign["date_to"]

    # 수수료율 조회 (마스터시트 J열)
    comm_rate = 0.0
    try:
        creds  = Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=SCOPES)
        cl     = gspread.authorize(creds)
        sh_id  = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", MASTER_SHEET_URL).group(1)
        rows   = cl.open_by_key(sh_id).sheet1.get_all_values()
        for r in rows[1:]:
            if len(r) > 9 and r[1].strip() == title:
                val = str(r[9]).replace(",", "").replace("%", "").strip()
                if val:
                    n = float(val)
                    comm_rate = n / 100 if n > 1 else n
                break
    except Exception:
        pass
    if not comm_rate:
        print(f"  [파마브로스 정산서 PDF] '{title[:30]}' 수수료율 조회 실패, 스킵")
        return

    settlement = round(total_payment * comm_rate)

    # 제품명 조회 (캠페인 실적 탭)
    product_name = ""
    try:
        creds2 = Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=SCOPES)
        cl2    = gspread.authorize(creds2)
        sh_id2 = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", MASTER_SHEET_URL).group(1)
        ws2    = cl2.open_by_key(sh_id2).worksheet("캠페인 실적(자사확인용)")
        for r in ws2.get_all_values()[1:]:
            if len(r) > 2 and r[1].strip() == title:
                product_name = r[2].strip()
                break
    except Exception:
        pass

    today = datetime.now(timezone(timedelta(hours=9))).strftime("%Y년 %m월 %d일")
    comm_pct = comm_rate * 100

    html = f"""<!DOCTYPE html>
<html lang="ko"><head><meta charset="UTF-8"><title>정산서</title><style>
@page{{size:A4;margin:12mm 15mm}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:"Malgun Gothic","맑은 고딕",sans-serif;background:#fff;color:#1a1a2e;max-width:680px;margin:0 auto;padding:0}}
.banner{{background:linear-gradient(135deg,#1a2744 0%,#2d3f6b 100%);padding:32px 40px 28px;position:relative;overflow:hidden}}
.banner::after{{content:"";position:absolute;right:-20px;top:-20px;width:120px;height:120px;border-radius:50%;background:rgba(255,255,255,0.05)}}
.banner-title{{color:#ffffff;font-size:30px;letter-spacing:6px;font-weight:700;margin-bottom:6px;text-shadow:0 1px 3px rgba(0,0,0,0.3)}}
.banner-sub{{color:rgba(255,255,255,0.9);font-size:13px;letter-spacing:2px;font-weight:500}}
.banner-date{{color:rgba(255,255,255,0.8);font-size:12px;text-align:right;margin-top:10px}}
.body{{padding:28px 40px 32px}}
.confirm-box{{background:#f0f4ff;border-left:4px solid #1a2744;padding:12px 16px;font-size:14px;color:#333;margin-bottom:24px;line-height:1.6}}
table{{width:100%;border-collapse:collapse;margin-bottom:6px}}
.tbl-head td{{background:#1a2744;color:#fff;padding:11px 16px;font-size:13px;font-weight:600;letter-spacing:0.5px}}
tr td{{padding:11px 16px;border-bottom:1px solid #eaecf4;font-size:14px}}
tr:nth-child(even) td{{background:#f8f9fb}}
tr td:first-child{{color:#555;width:38%}}
.total-row td{{font-weight:700;font-size:15px;color:#1a2744;background:#eef1fa!important;border-top:2px solid #1a2744}}
.note{{font-size:11.5px;color:#888;margin-top:8px;margin-bottom:24px;padding-left:2px}}
.company{{border:1px solid #dde1ee;border-radius:8px;padding:18px 22px;background:#fafbff}}
.company-title{{font-size:13px;font-weight:700;color:#1a2744;margin-bottom:10px;display:flex;align-items:center;gap:6px}}
.company-title::before{{content:"";display:inline-block;width:4px;height:14px;background:#1a2744;border-radius:2px}}
.company-body{{font-size:13px;color:#555;line-height:2}}
</style></head><body>
<div class="banner">
<div class="banner-title">정 산 서</div>
<div class="banner-sub">SETTLEMENT STATEMENT (파마브로스 양식)</div>
<div class="banner-date">발행일: {today}</div>
</div>
<div class="body">
<div class="confirm-box">공구진행에 따른 정산내역을 확인합니다.</div>
<table>
<tr class="tbl-head"><td colspan="2">정산 내역</td></tr>
<tr><td>인플루언서</td><td>{title}</td></tr>
{"<tr><td>제품명</td><td>" + product_name + "</td></tr>" if product_name else ""}
<tr><td>진행기간</td><td>{date_from} ~ {date_to}</td></tr>
<tr><td>총 결제금액</td><td>{total_payment:,}원</td></tr>
<tr><td>수수료</td><td>{comm_pct:.1f}%</td></tr>
<tr class="total-row"><td>정산기준금액</td><td>{settlement:,}원</td></tr>
</table>
<p class="note">* 세금관련부분은 협의된 내용으로 처리됩니다. (ex) 부가세 포함, 프리랜서 공제 등</p>
<div class="company">
<div class="company-title">정산 업체 정보</div>
<div class="company-body">업체명&nbsp;&nbsp;&nbsp;주식회사 정담건강<br>사업자번호&nbsp;&nbsp;&nbsp;391-86-00889<br>주소&nbsp;&nbsp;&nbsp;경기도 시흥시 서울대학로278번길61, 431-2호</div>
</div>
</div>
</body></html>"""

    try:
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaInMemoryUpload
        from google.oauth2.credentials import Credentials as OAuthCreds
        from google.auth.transport.requests import Request as GRequest

        oauth_creds = OAuthCreds(
            token=None,
            refresh_token=PHARMABROS_OAUTH_REFRESH_TOKEN,
            token_uri="https://oauth2.googleapis.com/token",
            client_id=PHARMABROS_OAUTH_CLIENT_ID,
            client_secret=PHARMABROS_OAUTH_CLIENT_SECRET,
        )
        oauth_creds.refresh(GRequest())
        svc = build("drive", "v3", credentials=oauth_creds, cache_discovery=False)

        # 파일명 결정 (완료 폴더 xlsx 기반)
        import re as _re
        safe_title = _re.sub(r'[\\/:*?"<>|]', "_", title).strip()
        pdf_name = f"{safe_title}_정산서.pdf"
        try:
            done_q = (
                f"name='완료' and mimeType='application/vnd.google-apps.folder' "
                f"and '{PHARMABROS_DRIVE_FOLDER_ID}' in parents and trashed=false"
            )
            done_folders = svc.files().list(q=done_q, fields="files(id)").execute().get("files", [])
            if done_folders:
                done_fid = done_folders[0]["id"]
                xlsx_files = svc.files().list(
                    q=f"'{done_fid}' in parents and trashed=false",
                    fields="files(name)",
                    orderBy="createdTime desc",
                ).execute().get("files", [])
                for xf in xlsx_files:
                    if xf["name"].startswith(safe_title + "_") and xf["name"].endswith(".xlsx"):
                        base = xf["name"][:-5]          # .xlsx 제거
                        base = base.replace("_업로드건", "")
                        pdf_name = base + "_정산서.pdf"
                        break
        except Exception:
            pass

        # '정산서' 폴더 내 동일 캠페인 기존 파일 전부 삭제 (중복 방지)
        try:
            old_files = svc.files().list(
                q=f"'{PHARMABROS_SETTLEMENT_FOLDER_ID}' in parents and trashed=false",
                fields="files(id,name)",
            ).execute().get("files", [])
            for of in old_files:
                if (of["name"].startswith(safe_title + "_")
                        or of["name"] in (f"{title} 정산서.pdf", f"{title} 정산서.html")):
                    svc.files().update(fileId=of["id"], body={"trashed": True}).execute()
                    print(f"  [파마브로스 정산서] 기존 파일 삭제: {of['name']}")
        except Exception:
            pass

        # playwright로 HTML → PDF 변환 (CSS gradient/border 등 완전 보존)
        from playwright.sync_api import sync_playwright
        import base64 as _b64
        data_uri = "data:text/html;charset=utf-8;base64," + _b64.b64encode(html.encode("utf-8")).decode()
        with sync_playwright() as pw:
            browser = pw.chromium.launch(args=["--no-sandbox", "--disable-setuid-sandbox"])
            page = browser.new_page()
            page.goto(data_uri, wait_until="networkidle")
            pdf_bytes = page.pdf(
                format="A4",
                print_background=True,
                margin={"top": "10mm", "bottom": "10mm", "left": "15mm", "right": "15mm"},
            )
            browser.close()

        pdf_media = MediaInMemoryUpload(pdf_bytes, mimetype="application/pdf", resumable=False)
        svc.files().create(
            body={"name": pdf_name, "parents": [PHARMABROS_SETTLEMENT_FOLDER_ID]},
            media_body=pdf_media,
        ).execute()

        print(f"  [파마브로스 정산서] '{title[:30]}' → {pdf_name} 업로드 완료")

    except Exception as e:
        print(f"  [파마브로스 정산서 오류] {e}")


# ── 사은품 추첨 ────────────────────────────────────────────
SAEUNPUM_TAB_NAME = "사은품당첨자"
SAEUNPUM_HEADER  = ["", "상품코드", "상품명", "수량", "수령자명", "전화번호", "주소", "메모", "내용", "전화끝4자리"]


def _get_or_create_saeunpum_ws(ss):
    """마스터시트에서 '사은품당첨자' 탭을 반환하거나 신규 생성한다."""
    try:
        return ss.worksheet(SAEUNPUM_TAB_NAME)
    except gspread.WorksheetNotFound:
        ws = ss.add_worksheet(title=SAEUNPUM_TAB_NAME, rows=1000, cols=len(SAEUNPUM_HEADER))
        ws.append_row(SAEUNPUM_HEADER, value_input_option="USER_ENTERED")
        return ws


def _run_saeunpum_lottery(force: bool = False):
    """
    마스터시트 시트1 L열(사은품인원)에 숫자가 있는 종료 캠페인을 찾아 추첨 후
    마스터시트의 '사은품당첨자' 탭에 최신순(위)으로 기재한다.

    날짜 조건 (force=False일 때):
      종료일 +1일 ≤ today ≤ 종료일 +7일
    force=True → 날짜 체크 없이 즉시 실행 (테스트용)
    """
    import random

    print("\n[사은품 추첨] 시작...")

    # ── 마스터시트 읽기 ──────────────────────────────────────
    if not MASTER_SHEET_URL:
        print("  [사은품] MASTER_SHEET_URL 없음")
        return

    try:
        creds = Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=SCOPES)
        gc = gspread.authorize(creds)
        sh_id = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", MASTER_SHEET_URL).group(1)
        ss = gc.open_by_key(sh_id)
        ws1 = ss.sheet1  # 시트1
        master_rows = ws1.get_all_records()
    except Exception as e:
        print(f"  [사은품] 마스터시트 읽기 실패: {e}")
        return

    # ── 사은품당첨자 탭 (마스터시트 내) ─────────────────────
    try:
        saeunpum_ws = _get_or_create_saeunpum_ws(ss)
        existing_rows = saeunpum_ws.get_all_values()
        # I열(index 8) 기재된 내용 목록 (중복 방지)
        existing_i = {r[8] for r in existing_rows[1:] if len(r) > 8 and r[8].strip()}
    except Exception as e:
        print(f"  [사은품] 사은품당첨자 탭 읽기 실패: {e}")
        return

    KST = timezone(timedelta(hours=9))
    today = datetime.now(KST).date()
    processed = 0

    for row in master_rows:
        title     = str(row.get("제목") or "").strip()
        start_str = str(row.get("시작일자") or "").strip()
        end_str   = str(row.get("종료일자") or "").strip()
        url       = str(row.get("상품링크") or "").strip()
        store     = str(row.get("스토어") or "").strip().lower()

        # L열: 헤더명이 길어서 포함 검색
        saeunpum_key = next(
            (k for k in row.keys() if "사은품인원" in k), None
        )
        saeunpum_n_raw = str(row.get(saeunpum_key) or "").strip() if saeunpum_key else ""

        if not title or not saeunpum_n_raw:
            continue

        try:
            saeunpum_n = int(saeunpum_n_raw)
        except ValueError:
            continue

        if saeunpum_n <= 0:
            continue

        # 날짜 파싱
        try:
            start_date = parse_date(start_str)
            end_date   = parse_date(end_str)
        except Exception:
            continue

        # 날짜 조건 체크 (업체가 사은품처리를 늦게 요청하는 경우 고려 → 최대 30일)
        days_since_end = (today - end_date).days
        if not force and (days_since_end < 1 or days_since_end > 30):
            continue

        # 종료 바로 다음날(days_since_end==1)은 오전 7시 이후에만 처리
        if not force and days_since_end == 1 and datetime.now(KST).hour < 7:
            print(f"  [사은품] '{title[:30]}' 종료 다음날이지만 오전 7시 이전 → 스킵")
            continue

        # 이미 처리된 캠페인 스킵 (I열에 제목이 포함된 행이 있으면)
        already = any(title in ic for ic in existing_i)
        if already:
            print(f"  [사은품] '{title[:30]}' 이미 처리됨, 스킵")
            continue

        # 스토어 API 자격증명
        if store not in STORE_CREDENTIALS:
            print(f"  [사은품] '{title[:30]}' 스토어 '{store}' 미등록")
            continue
        api_id, api_secret = STORE_CREDENTIALS[store]
        if not api_id or not api_secret:
            print(f"  [사은품] '{title[:30]}' API 키 없음")
            continue

        # 상품번호 추출
        try:
            product_no = extract_product_no(url)
        except ValueError:
            print(f"  [사은품] '{title[:30]}' 상품번호 추출 실패")
            continue

        # I열 내용 포맷: "2026.7.21~27_띱약사_상품명"
        # 종료월이 시작월과 다를 경우(월 넘어감) 월.일 형식으로 표시
        start_fmt = f"{start_date.year}.{start_date.month}.{start_date.day}"
        if end_date.month == start_date.month:
            end_part = str(end_date.day)
        else:
            end_part = f"{end_date.month}.{end_date.day}"
        i_col_prefix = f"{start_fmt}~{end_part}_{title}_"

        print(f"\n  [사은품] '{title[:30]}' → {saeunpum_n}명 추첨 시작...")

        # ── Naver API로 주문 + 구매자 정보 조회 ──────────────
        try:
            orders, product_name = naver_api.get_saeunpum_orders(
                client_id=api_id,
                client_secret=api_secret,
                product_no=product_no,
                date_from=start_date.strftime("%Y-%m-%d"),
                date_to=end_date.strftime("%Y-%m-%d"),
            )
        except Exception as e:
            print(f"  [사은품] '{title[:30]}' 주문 조회 오류: {e}")
            continue

        if not orders:
            print(f"  [사은품] '{title[:30]}' 주문 없음")
            continue

        i_col = i_col_prefix + product_name

        # C열용 상품명: [인플루언서태그] 접두어 제거
        product_name_clean = re.sub(r'^\[.*?\]\s*', '', product_name).strip()

        # ── 구매자별 주문 그룹화 ─────────────────────────────
        buyer_map = {}  # (수령자명, 전화번호) → 첫 번째 주문 dict
        buyer_count = {}
        for o in orders:
            key = (o["수령자명"].strip(), o["전화번호"].strip())
            if key not in buyer_map:
                buyer_map[key] = o
            buyer_count[key] = buyer_count.get(key, 0) + 1

        all_buyers  = list(buyer_map.keys())
        multi_buyers = [k for k in all_buyers if buyer_count[k] >= 2]
        print(f"  → 전체 구매자: {len(all_buyers)}명, 다구매자(2건+): {len(multi_buyers)}명")

        # ── 추첨 (70% 다구매자, 30% 전체) ───────────────────
        n_multi = round(saeunpum_n * 0.7)

        if len(multi_buyers) <= n_multi:
            winners_multi = multi_buyers[:]
        else:
            winners_multi = random.sample(multi_buyers, n_multi)

        # 다구매자 부족분을 전체 풀에서 채워 항상 saeunpum_n명 확보
        n_rest = saeunpum_n - len(winners_multi)
        already_won = set(winners_multi)
        rest_pool   = [k for k in all_buyers if k not in already_won]
        winners_rest = random.sample(rest_pool, min(n_rest, len(rest_pool)))

        winners = winners_multi + winners_rest
        print(f"  → 당첨: 다구매자 {len(winners_multi)}명 + 일반 {len(winners_rest)}명 = 총 {len(winners)}명")

        # ── 사은품당첨자 탭에 기재 (최신순: 헤더 바로 아래 삽입) ─
        new_rows = []
        for key in winners:
            o = buyer_map[key]
            tel = o["전화번호"]
            tel_last4 = tel[-4:] if len(tel) >= 4 else tel
            # [A, B, C, D, E, F, G, H, I, J]
            new_rows.append([
                "",                  # A: pimz_order_id (공백)
                "",                  # B: 상품코드 (이지어드민, 불가)
                product_name_clean,  # C: 상품명 ([태그] 제거)
                1,                   # D: 수량 (고정)
                o["수령자명"], # E: 수령자명
                o["전화번호"], # F: 전화번호
                o["주소"],    # G: 주소
                "",            # H: 메모
                i_col,         # I: 내용
                tel_last4,     # J: 전화끝4자리
            ])

        if new_rows:
            try:
                # row=2 삽입 → 헤더(1행) 바로 아래, 최신이 항상 위
                saeunpum_ws.insert_rows(new_rows, row=2, value_input_option="USER_ENTERED")
                print(f"  → 사은품당첨자 탭에 {len(new_rows)}행 기재 완료 (최신순)")
                existing_i.add(i_col)  # 중복 방지 업데이트
                processed += 1
            except Exception as e:
                print(f"  [사은품] 탭 기재 오류: {e}")

    if processed == 0:
        print("  [사은품 추첨] 처리할 캠페인 없음 (날짜 조건 미충족 or 이미 처리됨)")
    else:
        print(f"  [사은품 추첨] 완료: {processed}개 캠페인 처리")


def _run_pharmabros_if_needed(force: bool = False):
    """
    파마브로스파일공유여부=파마브로스파일공유 캠페인을 찾아,
    현재 시각이 업로드 시간대(KST 10시/14시/16시)이면 실행.
    force=True → 시간 체크 없이 즉시 실행 (테스트용)
    """
    global PHARMABROS_DRIVE_FOLDER_ID

    pb_campaigns = load_pharmabros_campaigns()
    if not pb_campaigns:
        print("  [파마브로스] 대상 캠페인 없음 (K열 확인 필요)")
        return

    print(f"\n  [파마브로스] 대상 캠페인 {len(pb_campaigns)}개 확인")

    for campaign in pb_campaigns:
        title = campaign["title"]

        # ── 자동 삭제 대상일 처리 ────────────────────────────
        if campaign.get("is_delete_day"):
            now_h = datetime.now(timezone(timedelta(hours=9))).hour
            if not force and now_h != 10:
                print(f"  [파마브로스] '{title[:30]}' — 삭제 예정일이나 10시 아님, 스킵")
                continue
            try:
                deleted = pharmabros.delete_campaign_files(
                    client_id=PHARMABROS_OAUTH_CLIENT_ID,
                    client_secret=PHARMABROS_OAUTH_CLIENT_SECRET,
                    refresh_token=PHARMABROS_OAUTH_REFRESH_TOKEN,
                    folder_id=PHARMABROS_DRIVE_FOLDER_ID,
                    title=title,
                )
                if deleted:
                    print(f"  [파마브로스] 📁 '{title[:30]}' 완료 폴더 이동 완료: {deleted}")
                else:
                    print(f"  [파마브로스] '{title[:30]}' 이동할 파일 없음 (이미 이동됨)")
            except Exception as e:
                print(f"  [파마브로스 삭제 오류] {e}")

            # 완료 이동과 동시에 정산서 PDF 자동 업로드
            try:
                total_payment, _ = _calc_pharmabros_settlement(campaign)
                _upload_pharmabros_settlement_pdf(campaign, total_payment)
            except Exception as e:
                print(f"  [파마브로스 정산서 PDF 오류] {e}")

            continue  # 업로드 로직은 실행하지 않음
        start_date = campaign["date_from"]
        end_date   = campaign["date_to"]
        is_final   = campaign["is_final"]

        if not force:
            run_flag, _ = pharmabros.should_run(start_date, end_date)
            if not run_flag:
                print(f"  [파마브로스] '{title[:30]}' — 현재 시각은 업로드 시간대 아님, 스킵")
                continue
        else:
            print(f"  [파마브로스] '{title[:30]}' — 강제 실행 모드 (시간 체크 무시)")

        try:
            # OAuth2 토큰 + 폴더 ID 확인
            if not all([PHARMABROS_OAUTH_CLIENT_ID,
                        PHARMABROS_OAUTH_CLIENT_SECRET,
                        PHARMABROS_OAUTH_REFRESH_TOKEN,
                        PHARMABROS_DRIVE_FOLDER_ID]):
                print(
                    "  [파마브로스] ⚠️  .env에 OAuth2 설정이 없습니다.\n"
                    "  → get_drive_token.py 를 실행해서 토큰을 발급받으세요."
                )
                continue  # 기존 기능에 영향 없이 스킵

            pharmabros.run_pharmabros(
                campaign=campaign,
                credentials_path=CREDENTIALS_PATH,
                oauth_client_id=PHARMABROS_OAUTH_CLIENT_ID,
                oauth_client_secret=PHARMABROS_OAUTH_CLIENT_SECRET,
                oauth_refresh_token=PHARMABROS_OAUTH_REFRESH_TOKEN,
                drive_folder_id=PHARMABROS_DRIVE_FOLDER_ID,
            )

            # ── 캠페인 중간 시점 이후: 옵션가격 자동 저장 ────────
            _save_pharmabros_option_prices(campaign)

            # ── 최종 업로드 시: I열 매출 + 정산탭 내역 기재
            if is_final:
                _write_pharmabros_revenue_to_i(campaign)
                settlement, breakdown = _calc_pharmabros_settlement(campaign)
                if settlement > 0:
                    _write_pharmabros_settlement(title, settlement, breakdown)

        except Exception as e:
            print(f"  [파마브로스 오류] {e}")
            err_str = str(e)
            if "invalid_grant" in err_str:
                cause  = "구글 인증 토큰 만료 또는 취소됨"
                action = (
                    "조치 방법:\n"
                    "1. 사장님 PC에서 python get_drive_token.py 실행\n"
                    "2. 새 OAUTH_REFRESH_TOKEN 복사\n"
                    "3. SSH 서버에서 아래 명령어 실행:\n"
                    "sed -i \"s|OAUTH_REFRESH_TOKEN=.*|OAUTH_REFRESH_TOKEN=새토큰값|\" "
                    "~/influencer-sales-data-share/'influencer data shared'/.env"
                )
            elif "503" in err_str or "Transient" in err_str or "transient" in err_str:
                cause  = "구글 서버 일시적 오류 (503) — 자동 재시도 예정"
                action = "별도 조치 불필요. 5~10분 후 자동으로 재시도됩니다."
            elif "403" in err_str or "Forbidden" in err_str:
                cause  = "구글 드라이브 접근 권한 오류"
                action = "드라이브 폴더 공유 설정 또는 OAuth 앱 권한을 확인해주세요."
            elif "quota" in err_str.lower() or "429" in err_str:
                cause  = "구글 API 할당량 초과"
                action = "잠시 후 자동 재시도됩니다. 반복 시 문의해주세요."
            else:
                cause  = f"알 수 없는 오류: {err_str[:120]}"
                action = "개발자에게 문의해주세요."
            send_telegram(
                f"⚠️ [파마브로스 파일공유 오류]\n\n"
                f"📦 캠페인: {title[:40]}\n"
                f"❌ 원인: {cause}\n\n"
                f"🔧 {action}\n\n"
                f"🕐 {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            )


def main():
    # 이익 2/3 조정: 캠페인 실적 탭 N, O열 기존값을 2/3 숫자로 즉시 변환
    if "--fix-profit" in sys.argv:
        print("\n[이익 2/3 조정] 캠페인 실적 탭 N, O열 변환 중...\n")
        sheets.fix_profit_columns_twothird(MASTER_SHEET_URL)
        print("완료")
        return

    # 파마브로스 정산서 PDF 즉시 업로드: python3 main.py --upload-pdf
    if "--upload-pdf" in sys.argv:
        print("\n[파마브로스 정산서 PDF] 종료된 캠페인 정산서 Drive 업로드 중...\n")
        try:
            creds   = Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=SCOPES)
            client  = gspread.authorize(creds)
            sh_id   = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", MASTER_SHEET_URL).group(1)
            ws      = client.open_by_key(sh_id).sheet1
            rows    = ws.get_all_records()
            KST     = timezone(timedelta(hours=9))
            today   = datetime.now(KST).date()
            uploaded = 0
            for row in rows:
                k_val = re.sub(r"\s+", "", str(row.get("파마브로스파일공유여부", "")))
                if k_val != "파마브로스파일공유":
                    continue
                title     = str(row.get("제목", "")).strip()
                end_str   = str(row.get("종료일자", "")).strip()
                start_str = str(row.get("시작일자", "")).strip()
                url       = str(row.get("상품링크", "")).strip()
                store     = str(row.get("스토어", "")).strip().lower()
                if not all([title, end_str, start_str, url, store]):
                    continue
                if store not in STORE_CREDENTIALS:
                    continue
                try:
                    end_date = parse_date(end_str)
                except ValueError:
                    continue
                if end_date >= today:
                    continue
                api_id, api_secret = STORE_CREDENTIALS[store]
                try:
                    product_no = extract_product_no(url)
                except ValueError:
                    continue
                campaign = {
                    "title": title, "product_no": product_no, "url": url,
                    "date_from": parse_date(start_str).strftime("%Y-%m-%d"),
                    "date_to": end_date.strftime("%Y-%m-%d"),
                    "api_id": api_id, "api_secret": api_secret, "store": store,
                }
                print(f"  처리 중: {title[:30]}")
                total_payment, _ = _calc_pharmabros_settlement(campaign)
                _upload_pharmabros_settlement_pdf(campaign, total_payment)
                uploaded += 1
        except Exception as e:
            print(f"  오류: {e}")
        print(f"\n완료: {uploaded}개 업로드")
        return

    # 파마브로스 주문상태 갱신: python3 main.py --refresh-status
    # 완료폴더 xlsx의 주문상태 컬럼만 네이버 API로 재조회해서 업데이트 (나머지 컬럼 절대 변경 안 함)
    if "--refresh-status" in sys.argv:
        print("\n[파마브로스 주문상태 갱신] 완료폴더 xlsx 주문상태 재조회 중...\n")
        if not all([PHARMABROS_OAUTH_CLIENT_ID, PHARMABROS_OAUTH_CLIENT_SECRET,
                    PHARMABROS_OAUTH_REFRESH_TOKEN, PHARMABROS_DRIVE_FOLDER_ID]):
            print("  ⚠️  OAuth2 설정 없음")
            return
        try:
            from googleapiclient.http import MediaIoBaseUpload as _MediaUpload
            svc = pharmabros._drive_service_oauth(
                PHARMABROS_OAUTH_CLIENT_ID,
                PHARMABROS_OAUTH_CLIENT_SECRET,
                PHARMABROS_OAUTH_REFRESH_TOKEN,
            )
            # 완료 폴더 ID 조회
            done_folders = svc.files().list(
                q=f"'{PHARMABROS_DRIVE_FOLDER_ID}' in parents and name='완료' and "
                  "mimeType='application/vnd.google-apps.folder' and trashed=false",
                fields="files(id)",
            ).execute().get("files", [])
            if not done_folders:
                print("  ⚠️  완료 폴더 없음")
                return
            done_folder_id = done_folders[0]["id"]

            creds   = Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=SCOPES)
            client  = gspread.authorize(creds)
            sh_id   = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", MASTER_SHEET_URL).group(1)
            rows    = client.open_by_key(sh_id).sheet1.get_all_records()
            KST     = timezone(timedelta(hours=9))
            today   = datetime.now(KST).date()
            refreshed = 0

            for row in rows:
                k_val = re.sub(r"\s+", "", str(row.get("파마브로스파일공유여부", "")))
                if k_val != "파마브로스파일공유":
                    continue
                title     = str(row.get("제목", "")).strip()
                end_str   = str(row.get("종료일자", "")).strip()
                store     = str(row.get("스토어", "")).strip().lower()
                url       = str(row.get("상품링크", "")).strip()
                if not all([title, end_str, store, url]) or store not in STORE_CREDENTIALS:
                    continue
                try:
                    end_date = parse_date(end_str)
                except ValueError:
                    continue
                # 종료 후 10~20일 사이에만 처리 (이후엔 영구 스킵)
                days_after = (today - end_date).days
                if days_after < 10:
                    print(f"  스킵 (종료 {days_after}일, 10일 미만): {title[:30]}")
                    continue
                if days_after > 20:
                    continue
                api_id, api_secret = STORE_CREDENTIALS[store]
                if not api_id or not api_secret:
                    continue

                safe_prefix = re.sub(r'[\\/:*?"<>|]', "_", title).strip() + "_"
                # 완료 폴더에서 해당 파일 찾기
                files = svc.files().list(
                    q=f"'{done_folder_id}' in parents and trashed=false",
                    fields="files(id,name)",
                    pageSize=100,
                ).execute().get("files", [])
                target = next(
                    (f for f in files
                     if f["name"].startswith(safe_prefix) and f["name"].endswith(".xlsx")),
                    None,
                )
                if not target:
                    print(f"  파일 없음 (완료폴더): {title[:30]}")
                    continue

                print(f"  처리 중: {target['name']}")
                import io as _io
                from openpyxl import load_workbook as _load_wb
                content = svc.files().get_media(fileId=target["id"]).execute()
                wb = _load_wb(filename=_io.BytesIO(content))
                ws = wb.active

                # 헤더 위치 파악
                header_row_idx = col_oid = col_status = None
                for hrow in ws.iter_rows(min_row=1, max_row=10):
                    cells = [str(c.value or "").strip() for c in hrow]
                    if "주문번호" in cells:
                        header_row_idx = hrow[0].row
                        col_oid    = cells.index("주문번호") + 1
                        col_status = cells.index("주문상태") + 1 if "주문상태" in cells else None
                        break
                if not header_row_idx or not col_oid or not col_status:
                    print(f"    헤더 파악 실패, 스킵")
                    continue

                # 주문번호 수집
                order_ids  = []
                data_rows  = []
                for drow in ws.iter_rows(min_row=header_row_idx + 1):
                    oid = str(drow[col_oid - 1].value or "").strip()
                    if oid:
                        order_ids.append(oid)
                        data_rows.append(drow)

                # API 일괄 조회 (100개씩)
                status_map = {}
                token = naver_api._get_access_token(api_id, api_secret)
                headers_api = {"Authorization": f"Bearer {token}"}
                for start in range(0, len(order_ids), 100):
                    chunk = order_ids[start:start + 100]
                    resp  = _requests.post(
                        f"{naver_api.BASE_URL}/external/v1/pay-order/seller/product-orders/query",
                        headers={**headers_api, "Content-Type": "application/json"},
                        json={"productOrderIds": chunk},
                        timeout=30,
                    )
                    if resp.status_code != 200:
                        print(f"    query 실패: {resp.status_code}")
                        continue
                    for order in resp.json().get("data", []):
                        po  = order.get("productOrder", {})
                        oid = str(po.get("productOrderId", ""))
                        raw = po.get("productOrderStatus", "")
                        status_map[oid] = naver_api.ORDER_STATUS_KO.get(raw, raw)

                # 주문상태 컬럼만 업데이트 (절대 다른 컬럼 건드리지 않음)
                updated = 0
                for drow in data_rows:
                    oid  = str(drow[col_oid - 1].value or "").strip()
                    if oid not in status_map:
                        continue
                    cell = drow[col_status - 1]
                    if cell.value != status_map[oid]:
                        cell.value = status_map[oid]
                        updated += 1

                if updated == 0:
                    print(f"    변경 없음")
                    continue

                out = _io.BytesIO()
                wb.save(out)
                out.seek(0)
                svc.files().update(
                    fileId=target["id"],
                    media_body=_MediaUpload(
                        out,
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    ),
                ).execute()
                print(f"    ✅ {updated}건 업데이트 → Drive 업로드 완료")
                refreshed += 1

        except Exception as e:
            print(f"  오류: {e}")
            import traceback; traceback.print_exc()
        print(f"\n완료: {refreshed}개 파일 업데이트")
        return

    # 파마브로스 테스트 모드: 시간 체크 없이 파마브로스만 즉시 실행
    if "--test-pharmabros" in sys.argv:
        print("\n[테스트 모드] 파마브로스 파일공유 강제 실행\n")
        _run_pharmabros_if_needed(force=True)
        _backfill_pharmabros_settlements()
        return

    # 사은품 추첨: 날짜 조건 포함 (로컬 오전 5:30 스케줄러 전용)
    if "--lottery" in sys.argv:
        _run_saeunpum_lottery()
        return

    # 사은품 추첨 테스트: 날짜 체크 없이 즉시 실행
    if "--test-lottery" in sys.argv:
        print("\n[테스트 모드] 사은품 추첨 강제 실행 (날짜 체크 무시)\n")
        _run_saeunpum_lottery(force=True)
        return

    # 완료 폴더 이동 테스트: python3 main.py --test-move 캠페인제목
    if "--test-move" in sys.argv:
        idx   = sys.argv.index("--test-move")
        title = sys.argv[idx + 1] if idx + 1 < len(sys.argv) else ""
        if not title:
            print("사용법: python3 main.py --test-move 캠페인제목")
            return
        print(f"\n[테스트 모드] '{title}' 완료 폴더 이동 강제 실행\n")
        moved = pharmabros.delete_campaign_files(
            client_id=PHARMABROS_OAUTH_CLIENT_ID,
            client_secret=PHARMABROS_OAUTH_CLIENT_SECRET,
            refresh_token=PHARMABROS_OAUTH_REFRESH_TOKEN,
            folder_id=PHARMABROS_DRIVE_FOLDER_ID,
            title=title,
        )
        if moved:
            print(f"✅ 완료 폴더로 이동됨: {moved}")
        else:
            print(f"이동할 파일 없음 ('{title}_' 로 시작하는 파일이 메인 폴더에 없음)")
        return

    # GitHub Actions 또는 --once 플래그: 1회 실행 후 종료
    if "--once" in sys.argv or os.getenv("GITHUB_ACTIONS"):
        run_once()
        return

    # 로컬 PC: 시작 즉시 1회 + 3시간마다 반복
    print("인플루언서 판매 데이터 공유 프로그램 시작")
    print("3시간마다 자동 업데이트  |  종료: Ctrl+C\n")
    run_once()
    schedule.every(3).hours.do(run_once)
    while True:
        schedule.run_pending()
        time.sleep(60)


if __name__ == "__main__":
    main()
