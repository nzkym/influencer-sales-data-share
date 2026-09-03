"""
P데이터(Drive 완료 폴더 xlsx) → Q데이터(파마브로스정산 탭) 동기화
- 완료 폴더의 모든 xlsx를 읽어 파마브로스정산 탭에 추가
- 주문번호 기준 중복 방지
- 취소완료/반품 포함 전체 행 저장 (Apps Script에서 필터링)

실행:
  python p_to_q_sync.py          # 실제 동기화
  python p_to_q_sync.py --dry    # 미리보기만 (시트 미수정)
"""
import re
import os
import io
import sys
from pathlib import Path
from dotenv import load_dotenv
from google.oauth2.credentials import Credentials as OAuthCredentials
from google.auth.transport.requests import Request as AuthRequest
from googleapiclient.discovery import build
from openpyxl import load_workbook
import gspread
from google.oauth2.service_account import Credentials

BASE_DIR = Path(__file__).parent
load_dotenv(BASE_DIR / ".env")

# pharmabros_oauth_setup.py 실행 후 생성되는 토큰 파일 우선 사용
_TOKEN_FILE = BASE_DIR / "pharmabros_token.json"
if _TOKEN_FILE.exists():
    import json as _json
    _tok = _json.loads(_TOKEN_FILE.read_text(encoding="utf-8"))
    OAUTH_CLIENT_ID     = _tok.get("client_id", "")
    OAUTH_CLIENT_SECRET = _tok.get("client_secret", "")
    OAUTH_REFRESH_TOKEN = _tok.get("refresh_token", "")
else:
    OAUTH_CLIENT_ID     = os.getenv("OAUTH_CLIENT_ID", "")
    OAUTH_CLIENT_SECRET = os.getenv("OAUTH_CLIENT_SECRET", "")
    OAUTH_REFRESH_TOKEN = os.getenv("OAUTH_REFRESH_TOKEN", "")

DONE_FOLDER_ID      = os.getenv("PHARMABROS_DONE_FOLDER_ID", "")
MAIN_FOLDER_ID      = os.getenv("PHARMABROS_DRIVE_FOLDER_ID", "")
MASTER_SHEET_URL    = os.getenv("MASTER_SHEET_URL", "")
CREDENTIALS_PATH    = str(BASE_DIR / "credentials" / "google-credentials.json")

DRY_RUN = "--dry" in sys.argv


def _drive_service():
    creds = OAuthCredentials(
        token=None,
        refresh_token=OAUTH_REFRESH_TOKEN,
        client_id=OAUTH_CLIENT_ID,
        client_secret=OAUTH_CLIENT_SECRET,
        token_uri="https://oauth2.googleapis.com/token",
    )
    creds.refresh(AuthRequest())
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def _sheets_client():
    creds = Credentials.from_service_account_file(
        CREDENTIALS_PATH,
        scopes=["https://www.googleapis.com/auth/spreadsheets"],
    )
    return gspread.authorize(creds)


def _parse_xlsx(content: bytes):
    """xlsx 바이트 → (title, date_from, date_to, orders) 반환."""
    wb = load_workbook(filename=io.BytesIO(content), read_only=True)
    ws = wb.active

    title, date_from, date_to = "", "", ""
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        if not row or not row[0]:
            continue
        cell = str(row[0])
        # 날짜 범위: (2026-08-10 ~ 2026-08-16)
        m = re.search(r'\((\d{4}-\d{2}-\d{2})\s*~\s*(\d{4}-\d{2}-\d{2})\)', cell)
        if m:
            date_from, date_to = m.group(1), m.group(2)
        # 제목: "[최종] 마미약사 판매현황 ..." → "마미약사"
        t = re.sub(r'\[최종\]|\[중간\]', '', cell).strip()
        t = re.split(r'판매현황', t)[0].strip()
        if t:
            title = t
        if date_from:
            break

    # 헤더 행 찾기 (주문번호 포함 행)
    header_row_idx = None
    headers = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        cells = [str(c or "").strip() for c in row]
        if "주문번호" in cells:
            header_row_idx = i
            headers = cells
            break

    orders = []
    if header_row_idx:
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row or not row[0]:
                continue
            entry = {}
            for i, h in enumerate(headers):
                if i < len(row) and h:
                    entry[h] = row[i] if row[i] is not None else ""
            order_no = str(entry.get("주문번호", "")).strip()
            # 합계 행("총 주문수량") 제외
            if order_no and "주문수량" not in order_no:
                orders.append(entry)

    wb.close()
    return title, date_from, date_to, orders


def sync():
    if not OAUTH_REFRESH_TOKEN:
        print("OAuth 토큰이 없습니다. pharmabros_oauth_setup.py를 먼저 실행해주세요.")
        return
    if not MASTER_SHEET_URL:
        print("MASTER_SHEET_URL이 .env에 없습니다.")
        return

    print(f"{'[DRY RUN] ' if DRY_RUN else ''}P→Q 동기화 시작")

    drive = _drive_service()

    # 완료 폴더 ID 결정
    folder_id = DONE_FOLDER_ID
    if not folder_id and MAIN_FOLDER_ID:
        # 메인 폴더에서 '완료' 하위폴더 자동 탐색
        done_files = drive.files().list(
            q=f"'{MAIN_FOLDER_ID}' in parents and name='완료' and mimeType='application/vnd.google-apps.folder' and trashed=false",
            fields="files(id)",
        ).execute().get("files", [])
        if done_files:
            folder_id = done_files[0]["id"]
            print(f"'완료' 하위폴더 자동 탐색: {folder_id}")
    if not folder_id:
        print("완료 폴더를 찾을 수 없습니다. .env에 PHARMABROS_DONE_FOLDER_ID 또는 PHARMABROS_DRIVE_FOLDER_ID를 추가해주세요.")
        return

    # 완료 폴더의 xlsx 목록
    files = drive.files().list(
        q=f"'{folder_id}' in parents and trashed=false",
        fields="files(id,name,createdTime)",
        orderBy="createdTime asc",
        pageSize=200,
    ).execute().get("files", [])

    xlsx_files = [f for f in files if f["name"].lower().endswith(".xlsx")]
    print(f"완료 폴더 xlsx 파일 {len(xlsx_files)}개")

    if not xlsx_files:
        print("동기화할 파일 없음")
        return

    # 시트 연결
    gc = _sheets_client()
    sheet_id = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", MASTER_SHEET_URL).group(1)
    ss = gc.open_by_key(sheet_id)

    try:
        pb_ws = ss.worksheet("파마브로스정산")
    except gspread.exceptions.WorksheetNotFound:
        print("파마브로스정산 탭이 없습니다.")
        return

    existing_rows = pb_ws.get_all_values()
    # 기존 주문번호 수집 (구버전=col[1], 신버전=col[2] 구분)
    existing_order_nos = set()
    for row in existing_rows[1:]:
        if not row:
            continue
        col1 = str(row[1]).strip() if len(row) > 1 else ""
        is_new_fmt = bool(re.match(r'^\d{4}-\d{2}-\d{2}$', col1))
        if is_new_fmt:
            order_no = str(row[2]).strip() if len(row) > 2 else ""
        else:
            order_no = col1  # 구버전: col[1]이 주문번호
        if order_no:
            existing_order_nos.add(order_no)
    print(f"기존 주문번호 {len(existing_order_nos)}개")

    new_rows = []
    for f in xlsx_files:
        print(f"\n  [{f['name']}]")
        try:
            content = drive.files().get_media(fileId=f["id"]).execute()
            title, date_from, date_to, orders = _parse_xlsx(content)
        except Exception as e:
            print(f"  ⚠️ 읽기 실패: {e}")
            continue

        if not title or not date_from:
            print(f"  ⚠️ 제목/날짜 파싱 실패, 스킵")
            continue

        print(f"  제목={title}, 기간={date_from}~{date_to}, 총 {len(orders)}건")

        added = 0
        for o in orders:
            order_no = str(o.get("주문번호", "")).strip()
            if not order_no or order_no in existing_order_nos:
                continue

            qty = o.get("주문수량", 0)
            price = o.get("단가", 0)
            try:
                qty_num = int(float(str(qty).replace(",", "") or 0)) if qty not in (None, "") else 0
                price_num = int(float(str(price).replace(",", "") or 0)) if price not in (None, "") else 0
                total = qty_num * price_num
            except (ValueError, TypeError):
                qty_num, price_num, total = 0, 0, 0

            # 신버전 9열 포맷:
            # [제목, 시작일(YYYY-MM-DD), 주문번호, 주문일시, 주문상태, 옵션, 주문수량, 단가, 단가총합]
            new_rows.append([
                title,
                date_from,
                order_no,
                str(o.get("주문일시", "")),
                str(o.get("주문상태", "")),
                str(o.get("옵션", "")),
                qty_num,
                price_num,
                total,
            ])
            existing_order_nos.add(order_no)
            added += 1

        print(f"  → 신규 {added}건")

    print(f"\n총 {len(new_rows)}건 추가 예정")
    if new_rows and not DRY_RUN:
        pb_ws.append_rows(new_rows, value_input_option="RAW")
        print("✅ 파마브로스정산 탭 동기화 완료")
    elif DRY_RUN:
        print("[DRY RUN] 실제 쓰기 생략")
        for r in new_rows[:5]:
            print(f"  {r}")
        if len(new_rows) > 5:
            print(f"  ... 외 {len(new_rows)-5}건")


if __name__ == "__main__":
    sync()
